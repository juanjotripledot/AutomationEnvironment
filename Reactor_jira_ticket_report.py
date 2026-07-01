#!/usr/bin/env python3
"""
Jira Ticket Report Generator — v4 (GitHub Actions Edition)
────────────────────────────────────────────────────────────────────────────────
Changes vs v3:
  • Generates TWO Excel files:
    1. Reactor_jira_year_report.xlsx — raw JIRA data (as before)
    2. RCTR_Host_Metrics_Report.xlsx — analyzed metrics with charts (new)
  • Reads Jira credentials from Environment Variables (Secrets).
  • Outputs Excel files to the local working directory (for GitHub Artifacts).
────────────────────────────────────────────────────────────────────────────────
"""

import json
import re
import logging
import os
import time
from datetime import datetime, timedelta, timezone
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import defaultdict
from statistics import mean

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference


# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION — Read from GitHub Secrets (Environment Variables)
# ══════════════════════════════════════════════════════════════════════════════
JIRA_BASE_URL  = os.environ.get("JIRA_BASE_URL", "https://tripledotstudios.atlassian.net")
JIRA_EMAIL     = os.environ.get("JIRA_EMAIL")
JIRA_API_TOKEN = os.environ.get("JIRA_API_TOKEN")

# Output paths
DATA_REPORT_PATH     = "Reactor_jira_year_report.xlsx"
METRICS_REPORT_PATH  = "RCTR_Host_Metrics_Report.xlsx"

# ══════════════════════════════════════════════════════════════════════════════
# TUNABLES
# ══════════════════════════════════════════════════════════════════════════════
MAX_WORKERS         = 8
SEARCH_PAGE_SIZE    = 100
CHANGELOG_PAGE_SIZE = 100
MAX_RETRIES         = 5
TIMEOUT             = 45

# ══════════════════════════════════════════════════════════════════════════════
# DOMAIN CONSTANTS
# ══════════════════════════════════════════════════════════════════════════════
TIME_IN_DEV_STATUSES = {
    "In Progress",
    "Re-Opened",
    "Code Review",
    "Deployed on Feature branch",
    "In QA feature branch",
    "Verified on feature branch",
}

HEADERS = [
    "Key",
    "Summary",
    "Story points",
    "Components",
    "Labels",
    "Assignee",
    "Description",
    "Comments",
    "Changelog (JSON)",
    "Linked Items",
    "Parent",
    "Depends on",
    "Current Status",
    "Last Sprint",
    "Number of Sub-tasks",
    "Number of Bug sub-tasks",
    "Time in In Progress",
    "Time in Bug Fixing",
    "Time in Code Review",
    "Time in Merged",
    "Time in Ready For QA",
    "Time in In QA",
    "Time in dev",
    "Delivery Time",
    "Times entered Bug Fixing",
]

STATUS_HEADER_MAP = {
    "Time in In Progress": "In Progress",
    "Time in Bug Fixing": "Bug Fixing",
    "Time in Code Review": "Code Review",
    "Time in Merged": "Merged",
    "Time in Ready For QA": "Ready For QA",
    "Time in In QA": "In QA",
}

FIRST_STATUS_COL = HEADERS.index("Time in In Progress")

# ══════════════════════════════════════════════════════════════════════════════
# LOGGING
# ══════════════════════════════════════════════════════════════════════════════
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ══════════════════════════════════════════════════════════════════════════════
# HTTP SESSION
# ══════════════════════════════════════════════════════════════════════════════
SESSION = requests.Session()
if JIRA_EMAIL and JIRA_API_TOKEN:
    SESSION.auth = (JIRA_EMAIL, JIRA_API_TOKEN)
SESSION.headers.update({
    "Accept": "application/json",
    "Content-Type": "application/json",
})


def _get(url, params=None, attempt=0):
    """GET with exponential-backoff retry."""
    try:
        resp = SESSION.get(url, params=params, timeout=TIMEOUT)

        if resp.status_code == 429 or resp.status_code >= 500:
            if attempt < MAX_RETRIES:
                wait = max(int(resp.headers.get("Retry-After", 0)), 2 ** attempt)
                log.warning("HTTP %s — backing off %ds (attempt %d/%d)",
                            resp.status_code, wait, attempt + 1, MAX_RETRIES)
                time.sleep(wait)
                return _get(url, params, attempt + 1)

        resp.raise_for_status()
        return resp.json()

    except requests.RequestException as exc:
        if attempt < MAX_RETRIES:
            wait = 2 ** attempt
            log.warning("Network error (%s) — retrying in %ds (attempt %d/%d)",
                        exc, wait, attempt + 1, MAX_RETRIES)
            time.sleep(wait)
            return _get(url, params, attempt + 1)

        log.error("Permanent failure for %s: %s", url, exc)
        return None


# ══════════════════════════════════════════════════════════════════════════════
# WORKING-HOURS CALCULATOR
# ══════════════════════════════════════════════════════════════════════════════

def working_hours_between(start, end):
    """Return working hours (Mon-Fri) between two timezone-aware datetimes."""
    if end <= start:
        return 0.0

    total_secs = 0.0
    day = start.replace(hour=0, minute=0, second=0, microsecond=0)

    while day < end:
        next_day = day + timedelta(days=1)
        if day.weekday() < 5:
            seg_start = max(start, day)
            seg_end   = min(end, next_day)
            if seg_end > seg_start:
                total_secs += (seg_end - seg_start).total_seconds()
        day = next_day

    return total_secs / 3600.0


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 - FETCH ALL ISSUES
# ══════════════════════════════════════════════════════════════════════════════

def fetch_all_issues():
    """Page through every issue matching the JQL."""
    jql = (
        'project in ("Reactor") AND component in ("Host Unity Package","Forge Unity Template","Reactor . Forge") '
        'AND (LABELS NOT IN (NoKPI) OR LABELS = EMPTY) '
        'AND issuetype not in (Documentation,Epic,"Sub-task","Bug sub-task") '
        'AND updated >= -365d '
        'ORDER BY updated DESC'
    )

    url             = f"{JIRA_BASE_URL}/rest/api/3/search/jql"
    issues          = []
    next_page_token = None
    offset          = 0
    page_num        = 0

    log.info("=== STEP 1: Fetching all issues via /rest/api/3/search/jql ===")

    while True:
        params = {
            "jql":        jql,
            "maxResults": SEARCH_PAGE_SIZE,
            "fields":     "key",
        }

        if next_page_token:
            params["nextPageToken"] = next_page_token
        else:
            params["startAt"] = offset

        data = _get(url, params=params)

        if not data:
            log.error("Empty response on page %d — stopping.", page_num)
            break

        batch = data.get("issues", [])
        if not batch:
            log.info("  Empty batch on page %d — reached the end.", page_num)
            break

        issues.extend(batch)
        page_num += 1

        total          = data.get("total", 0)
        next_page_token = data.get("nextPageToken") or data.get("next_page_token")

        log.info("  page=%-3d  batch=%-4d  accumulated=%-5d  total_reported=%-5s  cursor=%s",
                 page_num, len(batch), len(issues),
                 str(total) if total else "?",
                 "yes" if next_page_token else "no")

        if next_page_token:
            continue

        if total and len(issues) >= total:
            break
        if len(batch) < SEARCH_PAGE_SIZE:
            break

        offset += len(batch)

    log.info("Total issues collected: %d", len(issues))
    return issues


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 - FETCH FULL CHANGELOG
# ══════════════════════════════════════════════════════════════════════════════

def fetch_changelog(key):
    url       = f"{JIRA_BASE_URL}/rest/api/3/issue/{key}/changelog"
    histories = []
    start     = 0

    while True:
        data = _get(url, params={"startAt": start, "maxResults": CHANGELOG_PAGE_SIZE})

        if not data:
            break

        batch = data.get("values", [])
        if not batch:
            break

        histories.extend(batch)
        total = data.get("total", 0)

        if len(histories) >= total:
            break

        start += len(batch)

    return histories


# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 - FETCH ISSUE FIELDS
# ══════════════════════════════════════════════════════════════════════════════

def fetch_issue_fields(key):
    url  = f"{JIRA_BASE_URL}/rest/api/3/issue/{key}"
    data = _get(url, params={"fields": "*all"})
    return (data or {}).get("fields", {})


# ══════════════════════════════════════════════════════════════════════════════
# ADF -> PLAIN TEXT
# ══════════════════════════════════════════════════════════════════════════════

def _adf_to_text(node):
    if not node:
        return ""
    if isinstance(node, str):
        return node
    if node.get("type") == "text":
        return node.get("text", "")
    parts = [_adf_to_text(c) for c in node.get("content", [])]
    return " ".join(p for p in parts if p)


# ══════════════════════════════════════════════════════════════════════════════
# COMMENTS -> JSON STRING
# ══════════════════════════════════════════════════════════════════════════════

def build_comments_json(fields):
    result = []
    try:
        for c in fields.get("comment", {}).get("comments", []):
            body_obj = c.get("body", {})
            text     = _adf_to_text(body_obj) if isinstance(body_obj, dict) else str(body_obj or "")
            result.append({
                "datetime": c.get("created", ""),
                "author":   (c.get("author") or {}).get("displayName", "Unknown"),
                "comment":  text,
            })
    except Exception as exc:
        log.warning("Comments parse error: %s", exc)
    return json.dumps(result, ensure_ascii=False)


# ══════════════════════════════════════════════════════════════════════════════
# CHANGELOG -> JSON STRING
# ══════════════════════════════════════════════════════════════════════════════

def build_changelog_json(histories):
    result = []
    for h in histories:
        result.append({
            "id":      h.get("id", ""),
            "created": h.get("created", ""),
            "author":  (h.get("author") or {}).get("displayName", "Unknown"),
            "items": [
                {
                    "field":      item.get("field", ""),
                    "fieldtype":  item.get("fieldtype", ""),
                    "from":       item.get("from", ""),
                    "fromString": item.get("fromString", ""),
                    "to":         item.get("to", ""),
                    "toString":   item.get("toString", ""),
                }
                for item in h.get("items", [])
            ],
        })
    return json.dumps(result, ensure_ascii=False)


# ══════════════════════════════════════════════════════════════════════════════
# STATUS METRICS
# ══════════════════════════════════════════════════════════════════════════════

def _parse_ts(ts_str):
    if not ts_str:
        return None
    try:
        s = ts_str.strip()
        if s.endswith("Z"):
            s = s[:-1] + "+00:00"
        elif len(s) > 5 and s[-5] in "+-" and ":" not in s[-5:]:
            s = s[:-2] + ":" + s[-2:]
        return datetime.fromisoformat(s).astimezone(timezone.utc)
    except (ValueError, TypeError):
        return None


def compute_status_metrics(histories):
    now         = datetime.now(timezone.utc)
    transitions = []

    for h in histories:
        ts = _parse_ts(h.get("created", ""))
        if ts is None:
            continue
        for item in h.get("items", []):
            if item.get("field") == "status":
                transitions.append({
                    "ts":   ts,
                    "from": item.get("fromString", ""),
                    "to":   item.get("toString", ""),
                })

    transitions.sort(key=lambda x: x["ts"])

    time_in           = {}
    first_in_progress = None
    first_verified    = None
    bug_fixing_count  = 0

    for i, t in enumerate(transitions):
        name      = t["to"]
        enter     = t["ts"]
        exit_time = transitions[i + 1]["ts"] if i + 1 < len(transitions) else now

        wh = working_hours_between(enter, exit_time)

        time_in[name] = time_in.get(name, 0.0) + wh

        if name == "In Progress" and first_in_progress is None:
            first_in_progress = enter
        if name == "Verified for production" and first_verified is None:
            first_verified = enter
        if name == "Bug Fixing":
            bug_fixing_count += 1

    time_in_dev = sum(time_in.get(s, 0.0) for s in TIME_IN_DEV_STATUSES)

    delivery_time = None
    if first_in_progress and first_verified and first_verified > first_in_progress:
        delivery_time = working_hours_between(first_in_progress, first_verified)

    return time_in, time_in_dev, delivery_time, bug_fixing_count


# ══════════════════════════════════════════════════════════════════════════════
# PROCESS ONE ISSUE
# ══════════════════════════════════════════════════════════════════════════════

def process_issue(basic_issue):
    key = basic_issue.get("key", "UNKNOWN")
    try:
        fields = fetch_issue_fields(key)
        if not fields:
            log.warning("No fields for %s — skipping", key)
            return None

        summary     = fields.get("summary", "")
        sp          = fields.get("customfield_10023")
        components  = ", ".join(c.get("name", "") for c in (fields.get("components") or []))
        labels      = ", ".join(fields.get("labels") or [])
        assignee    = (fields.get("assignee") or {}).get("displayName", "")
        desc_obj    = fields.get("description")
        description = _adf_to_text(desc_obj) if isinstance(desc_obj, dict) else (desc_obj or "")
        comments    = build_comments_json(fields)

        current_status = fields.get("status", {}).get("name", "")

        last_sprint_name = ""
        sprint_list = []
        for k, v in fields.items():
            if v and isinstance(v, list) and len(v) > 0:
                if isinstance(v[0], dict) and "startDate" in v[0] and "endDate" in v[0] and "name" in v[0]:
                    sprint_list.extend(v)
                elif isinstance(v[0], str) and "id=" in v[0] and "name=" in v[0]:
                    for sp_str in v:
                        match = re.search(r"name=([^,]+)", sp_str)
                        if match:
                            sprint_list.append({"name": match.group(1)})

        if sprint_list:
            last_sprint_name = sprint_list[-1].get("name", "")

        subtasks = fields.get("subtasks", [])
        num_subtasks = 0
        num_subbugs = 0
        for st in subtasks:
            st_type = st.get("fields", {}).get("issuetype", {}).get("name", "")
            if st_type.lower() == "sub-task":
                num_subtasks += 1
            elif st_type.lower() == "bug sub-task":
                num_subbugs += 1

        current_project_key = key.split("-")[0] if "-" in key else ""
        depends_on_projects = set()

        # Extract Parent
        parent_key = ""
        parent_field = fields.get("parent")
        if parent_field:
            parent_key = parent_field.get("key", "")
            if parent_key:
                parent_proj = parent_key.split("-")[0]
                if parent_proj and parent_proj != current_project_key:
                    depends_on_projects.add(parent_proj)

        # Extract Linked Items
        linked_items_list = []
        issue_links = fields.get("issuelinks", [])
        for link in issue_links:
            link_type_name = link.get("type", {}).get("name", "")
            if "inwardIssue" in link:
                linked_issue = link["inwardIssue"]
                direction = link.get("type", {}).get("inward", link_type_name)
            elif "outwardIssue" in link:
                linked_issue = link["outwardIssue"]
                direction = link.get("type", {}).get("outward", link_type_name)
            else:
                continue

            linked_key = linked_issue.get("key", "")
            linked_items_list.append({
                "relationship": direction,
                "key": linked_key,
                "summary": (linked_issue.get("fields") or {}).get("summary", ""),
                "status": ((linked_issue.get("fields") or {}).get("status") or {}).get("name", "")
            })
            if linked_key:
                linked_proj = linked_key.split("-")[0]
                if linked_proj and linked_proj != current_project_key:
                    depends_on_projects.add(linked_proj)

        linked_items_json = json.dumps(linked_items_list, ensure_ascii=False)
        depends_on_str = ", ".join(sorted(list(depends_on_projects)))

        histories      = fetch_changelog(key)
        changelog_json = build_changelog_json(histories)

        time_in, time_in_dev, delivery_time, bug_fixing_count = compute_status_metrics(histories)

        row = [
            key,
            summary,
            sp,
            components,
            labels,
            assignee,
            description,
            comments,
            changelog_json,
            linked_items_json,
            parent_key,
            depends_on_str,
            current_status,
            last_sprint_name,
            num_subtasks,
            num_subbugs,
        ]

        for header in HEADERS[FIRST_STATUS_COL: FIRST_STATUS_COL + len(STATUS_HEADER_MAP)]:
            val = time_in.get(STATUS_HEADER_MAP[header], 0.0)
            row.append(round(val, 2) if val else None)

        row.append(round(time_in_dev, 2)   if time_in_dev    else None)
        row.append(round(delivery_time, 2) if delivery_time  else None)
        row.append(bug_fixing_count        if bug_fixing_count   else None)

        return row

    except Exception as exc:
        log.error("Error processing %s: %s", key, exc, exc_info=True)
        return None


# ══════════════════════════════════════════════════════════════════════════════
# BUILD EXCEL WORKBOOK (Data Report)
# ══════════════════════════════════════════════════════════════════════════════

ILLEGAL_XML_CHARS_RE = re.compile(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]")

def clean_for_excel(val):
    if isinstance(val, str):
        return ILLEGAL_XML_CHARS_RE.sub("", val)
    return val

def build_data_workbook(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Jira Tickets"

    h_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    h_fill  = PatternFill("solid", start_color="1F4E79")
    h_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.append(HEADERS)
    for col in range(1, len(HEADERS) + 1):
        cell           = ws.cell(row=1, column=col)
        cell.font      = h_font
        cell.fill      = h_fill
        cell.alignment = h_align
    ws.row_dimensions[1].height = 32

    d_font     = Font(name="Arial", size=10)
    d_align    = Alignment(wrap_text=True, vertical="top")
    num_format = "#,##0.00"
    numeric_start_col = FIRST_STATUS_COL + 1

    for row_data in rows:
        cleaned_row = [clean_for_excel(v) for v in row_data]
        ws.append(cleaned_row)
        r = ws.max_row
        for col in range(1, len(HEADERS) + 1):
            cell           = ws.cell(row=r, column=col)
            cell.font      = d_font
            cell.alignment = d_align
            if col >= numeric_start_col:
                cell.number_format = num_format

    widths = {
        1: 16, 2: 42, 3: 12, 4: 22, 5: 22, 6: 22, 7: 55, 8: 55, 9: 60, 10: 60,
        11: 16, 12: 24, 13: 20, 14: 25, 15: 18, 16: 18,
    }
    for col in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(col, 20)

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = ws.dimensions

    return wb


# ══════════════════════════════════════════════════════════════════════════════
# BUILD METRICS REPORT (from raw data)
# ══════════════════════════════════════════════════════════════════════════════

def build_metrics_report_from_data(rows):
    """Generate RCTR Host Metrics Report from extracted data"""
    log.info("=== Building Metrics Report ===")

    # Parse data for metrics
    data = []
    for row in rows:
        key = row[0]
        pts = row[2]
        sprint = row[13]
        current_status = row[12]
        num_subtasks = row[14] or 0
        num_subbugs = row[15] or 0
        parent = row[10]  # Parent epic

        # Get changelog to compute lead time
        try:
            changelog_json = row[8]
            changelog = json.loads(changelog_json) if changelog_json else []
        except:
            changelog = []

        # Extract status transitions
        first_status_time = {}
        all_states = set()

        for change in changelog:
            created_str = change.get('created', '')

            for item in change.get('items', []):
                if item.get('field') == 'status':
                    to_status_str = item.get('toString')
                    all_states.add(to_status_str)

                    if to_status_str and to_status_str not in first_status_time:
                        try:
                            fixed_created = created_str[:-2] + ':' + created_str[-2:]
                            dt = datetime.fromisoformat(fixed_created)
                            first_status_time[to_status_str] = dt
                        except:
                            pass

        # Only include if has In Progress and Done and has story points
        if 'In Progress' not in first_status_time or 'Done' not in first_status_time:
            continue
        if not pts:
            continue

        try:
            pts_int = int(pts)
        except:
            continue

        lead_time_days = (first_status_time['Done'] - first_status_time['In Progress']).total_seconds() / (24*3600)

        def time_between(from_st, to_st):
            if from_st in first_status_time and to_st in first_status_time:
                dt = (first_status_time[to_st] - first_status_time[from_st]).total_seconds() / (24*3600)
                return round(dt, 2)
            return None

        sprints = [s.strip() for s in str(sprint).split(',')] if sprint else []
        last_sprint = sprints[-1] if sprints else sprint

        total_bugs = num_subbugs + num_subtasks

        ticket = {
            'key': key,
            'sprint': last_sprint,
            'pts': pts_int,
            'lead_time': round(lead_time_days, 2),
            'dev': time_between('In Progress', 'Code Review'),
            'cr': time_between('Code Review', 'Merged'),
            'merged': time_between('Merged', 'Ready for QA'),
            'rqa': time_between('Ready for QA', 'In QA'),
            'qa': time_between('In QA', 'Done'),
            'bugs': total_bugs,
            'states_visited': all_states,
            'has_in_qa': 'In QA' in all_states,
            'has_ready_qa': 'Ready for QA' in all_states,
            'epic': parent or 'NO_EPIC',
            'first_in_progress': first_status_time.get('In Progress'),
            'first_done': first_status_time.get('Done'),
        }

        data.append(ticket)

    log.info("  Processed %d tickets with full metrics", len(data))

    # Group data
    grouped = defaultdict(list)
    for d in data:
        key = (d['sprint'], d['pts'])
        grouped[key].append(d)

    all_sprints = sorted(set(d['sprint'] for d in data if 'host' in str(d['sprint']).lower()),
                        key=lambda x: str(x).lower())
    all_points = sorted(set(d['pts'] for d in data))
    last_sprint = all_sprints[-1] if all_sprints else None

    # Create workbook
    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    h_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    h_font = Font(bold=True, color="FFFFFF")
    bord = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    cent = Alignment(horizontal='center')

    # ========== Sub-bugs Analysis ==========
    ws_bugs = wb_out.create_sheet("1. Análisis Sub-bugs", 0)

    ws_bugs.cell(1, 1).value = "Sprint / Pts"
    ws_bugs.cell(1, 1).fill = h_fill
    ws_bugs.cell(1, 1).font = h_font

    for col_idx, pt in enumerate(all_points, 2):
        cell = ws_bugs.cell(1, col_idx)
        cell.value = f"{pt}pt"
        cell.fill = h_fill
        cell.font = h_font
        cell.border = bord
        cell.alignment = cent

    for row_idx, sprint in enumerate(all_sprints, 2):
        cell = ws_bugs.cell(row_idx, 1)
        cell.value = sprint
        cell.fill = PatternFill(start_color="E8F0F8", end_color="E8F0F8", fill_type="solid")
        cell.border = bord

        for col_idx, pt in enumerate(all_points, 2):
            combo = (sprint, pt)
            items = grouped[combo]

            cell = ws_bugs.cell(row_idx, col_idx)
            cell.border = bord
            cell.alignment = cent

            if items:
                bugs_list = [item['bugs'] for item in items]
                avg_bugs = round(mean(bugs_list), 2)
                cell.value = avg_bugs
                cell.number_format = '0.00'

    ws_bugs.column_dimensions['A'].width = 16
    for col in range(2, len(all_points) + 2):
        ws_bugs.column_dimensions[get_column_letter(col)].width = 12

    chart = BarChart()
    chart.title = "Promedio Sub-bugs"
    chart.y_axis.title = '# Sub-bugs'
    data_ref = Reference(ws_bugs, min_col=2, min_row=1, max_col=len(all_points)+1, max_row=len(all_sprints)+1)
    cat_ref = Reference(ws_bugs, min_col=1, min_row=2, max_row=len(all_sprints)+1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cat_ref)
    chart.height = 12
    chart.width = 22
    ws_bugs.add_chart(chart, "A" + str(len(all_sprints) + 4))

    # ========== Metrics Sheets ==========
    metrics_list = [
        ('2. Lead Time', 'lead_time'),
        ('3. Tiempo Dev', 'dev'),
        ('4. Tiempo Code Review', 'cr'),
        ('5. Tiempo Merged', 'merged'),
        ('6. Tiempo Ready QA', 'rqa'),
        ('7. Tiempo QA', 'qa'),
    ]

    for metric_name, metric_key in metrics_list:
        ws_m = wb_out.create_sheet(metric_name)

        ws_m.cell(1, 1).value = "Sprint"
        ws_m.cell(1, 1).fill = h_fill
        ws_m.cell(1, 1).font = h_font

        for c_idx, p_val in enumerate(all_points, 2):
            cell = ws_m.cell(1, c_idx)
            cell.value = f"{p_val}pt"
            cell.fill = h_fill
            cell.font = h_font
            cell.border = bord
            cell.alignment = cent

        for r_idx, sprint_val in enumerate(all_sprints, 2):
            cell = ws_m.cell(r_idx, 1)
            cell.value = sprint_val
            cell.fill = PatternFill(start_color="E8F0F8", end_color="E8F0F8", fill_type="solid")
            cell.border = bord

            for c_idx, p_val in enumerate(all_points, 2):
                combo = (sprint_val, p_val)
                items = grouped[combo]

                cell = ws_m.cell(r_idx, c_idx)
                cell.border = bord
                cell.alignment = cent

                if items:
                    values = [item[metric_key] for item in items if item[metric_key] is not None]
                    if values:
                        avg = round(mean(values), 2)
                        cell.value = avg
                        cell.number_format = '0.00'

        chart = BarChart()
        chart.title = metric_name
        chart.y_axis.title = 'Días'
        data_ref = Reference(ws_m, min_col=2, min_row=1, max_col=len(all_points)+1, max_row=len(all_sprints)+1)
        cat_ref = Reference(ws_m, min_col=1, min_row=2, max_row=len(all_sprints)+1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cat_ref)
        chart.height = 10
        chart.width = 20
        ws_m.add_chart(chart, "A" + str(len(all_sprints) + 4))

        for c in range(1, len(all_points) + 2):
            ws_m.column_dimensions[get_column_letter(c)].width = 12

    # ========== Epic Lead Time Analysis ==========
    log.info("  Computing epic lead times...")

    # Group tickets by epic and sprint
    epic_data = defaultdict(lambda: defaultdict(list))  # epic -> sprint -> [tickets]

    for t in data:
        epic = t['epic']
        sprint = t['sprint']
        epic_data[epic][sprint].append(t)

    # Calculate lead time per epic per sprint
    epic_metrics = []  # (epic, sprint, lead_time_days, num_tickets)

    for epic in sorted(epic_data.keys()):
        for sprint in sorted(epic_data[epic].keys()):
            tickets = epic_data[epic][sprint]

            # Epic lead time = from first In Progress to last Done
            first_in_prog_list = [t['first_in_progress'] for t in tickets if t['first_in_progress']]
            first_done_list = [t['first_done'] for t in tickets if t['first_done']]

            if first_in_prog_list and first_done_list:
                min_in_prog = min(first_in_prog_list)
                max_done = max(first_done_list)

                if max_done > min_in_prog:
                    epic_lead_time = (max_done - min_in_prog).total_seconds() / (24*3600)
                    epic_metrics.append({
                        'epic': epic,
                        'sprint': sprint,
                        'lead_time': round(epic_lead_time, 2),
                        'num_tickets': len(tickets),
                    })

    log.info(f"  Found {len(epic_metrics)} epic-sprint combinations")

    all_epics = sorted(set(m['epic'] for m in epic_metrics))
    all_sprints_epic = sorted(set(m['sprint'] for m in epic_metrics))

    # ========== Epic Lead Time Details Sheet ==========
    if epic_metrics:
        ws_epic_detail = wb_out.create_sheet("8. Epic Lead Time Detail")

        headers_epic = ['Epic', 'Sprint', 'Lead Time (días)', '# Tickets']
        for col, h in enumerate(headers_epic, 1):
            cell = ws_epic_detail.cell(1, col)
            cell.value = h
            cell.fill = h_fill
            cell.font = h_font
            cell.border = bord
            cell.alignment = cent

        r = 2
        for m in epic_metrics:
            ws_epic_detail.cell(r, 1).value = m['epic']
            ws_epic_detail.cell(r, 2).value = m['sprint']
            ws_epic_detail.cell(r, 3).value = m['lead_time']
            ws_epic_detail.cell(r, 4).value = m['num_tickets']

            for col in range(1, 5):
                ws_epic_detail.cell(r, col).border = bord
                if col > 1:
                    ws_epic_detail.cell(r, col).alignment = cent
                    if col == 3:
                        ws_epic_detail.cell(r, col).number_format = '0.00'
            r += 1

        ws_epic_detail.column_dimensions['A'].width = 20
        ws_epic_detail.column_dimensions['B'].width = 16
        ws_epic_detail.column_dimensions['C'].width = 16
        ws_epic_detail.column_dimensions['D'].width = 12

    # ========== Epic Lead Time Matrix (Sprint x Epic) ==========
    if epic_metrics and all_epics and all_sprints_epic:
        ws_epic_matrix = wb_out.create_sheet("9. Epic Lead Time Matrix")

        # Header
        ws_epic_matrix.cell(1, 1).value = "Sprint / Epic"
        ws_epic_matrix.cell(1, 1).fill = h_fill
        ws_epic_matrix.cell(1, 1).font = h_font

        for col_idx, epic in enumerate(all_epics, 2):
            cell = ws_epic_matrix.cell(1, col_idx)
            cell.value = epic[:30]  # Truncate long epic names
            cell.fill = h_fill
            cell.font = h_font
            cell.border = bord
            cell.alignment = cent

        # Data rows
        for row_idx, sprint in enumerate(all_sprints_epic, 2):
            cell = ws_epic_matrix.cell(row_idx, 1)
            cell.value = sprint
            cell.fill = PatternFill(start_color="E8F0F8", end_color="E8F0F8", fill_type="solid")
            cell.border = bord

            for col_idx, epic in enumerate(all_epics, 2):
                matching = [m for m in epic_metrics if m['epic'] == epic and m['sprint'] == sprint]

                cell = ws_epic_matrix.cell(row_idx, col_idx)
                cell.border = bord
                cell.alignment = cent

                if matching:
                    cell.value = matching[0]['lead_time']
                    cell.number_format = '0.00'

        for col in range(1, len(all_epics) + 2):
            ws_epic_matrix.column_dimensions[get_column_letter(col)].width = 16

        # Chart
        chart = BarChart()
        chart.title = "Lead Time por Epic"
        chart.y_axis.title = 'Días'
        data_ref = Reference(ws_epic_matrix, min_col=2, min_row=1, max_col=len(all_epics)+1, max_row=len(all_sprints_epic)+1)
        cat_ref = Reference(ws_epic_matrix, min_col=1, min_row=2, max_row=len(all_sprints_epic)+1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cat_ref)
        chart.height = 12
        chart.width = 22
        ws_epic_matrix.add_chart(chart, "A" + str(len(all_sprints_epic) + 4))

    # ========== Epic Summary ==========
    if epic_metrics:
        ws_epic_summary = wb_out.create_sheet("10. Epic Summary")

        ws_epic_summary.cell(1, 1).value = "RESUMEN DE EPICS"
        ws_epic_summary.cell(1, 1).font = Font(bold=True, size=12)

        # Overall average
        row = 3
        ws_epic_summary.cell(row, 1).value = "Lead Time Promedio (Todos los Epics)"
        ws_epic_summary.cell(row, 2).value = round(mean([m['lead_time'] for m in epic_metrics]), 2)
        ws_epic_summary.cell(row, 1).border = bord
        ws_epic_summary.cell(row, 2).border = bord
        ws_epic_summary.cell(row, 2).number_format = '0.00'
        row += 1

        # Per sprint summary
        row += 1
        ws_epic_summary.cell(row, 1).value = "Lead Time Promedio por Sprint"
        ws_epic_summary.cell(row, 1).font = Font(bold=True)
        row += 1

        ws_epic_summary.cell(row, 1).value = "Sprint"
        ws_epic_summary.cell(row, 2).value = "Lead Time Promedio"
        ws_epic_summary.cell(row, 1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        ws_epic_summary.cell(row, 2).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        row += 1

        for sprint in all_sprints_epic:
            sprint_metrics = [m['lead_time'] for m in epic_metrics if m['sprint'] == sprint]
            if sprint_metrics:
                ws_epic_summary.cell(row, 1).value = sprint
                ws_epic_summary.cell(row, 2).value = round(mean(sprint_metrics), 2)
                ws_epic_summary.cell(row, 2).number_format = '0.00'
                ws_epic_summary.cell(row, 1).border = bord
                ws_epic_summary.cell(row, 2).border = bord
                row += 1

        # Per epic summary
        row += 1
        ws_epic_summary.cell(row, 1).value = "Lead Time Promedio por Epic"
        ws_epic_summary.cell(row, 1).font = Font(bold=True)
        row += 1

        ws_epic_summary.cell(row, 1).value = "Epic"
        ws_epic_summary.cell(row, 2).value = "Lead Time Promedio"
        ws_epic_summary.cell(row, 3).value = "# Sprints"
        ws_epic_summary.cell(row, 1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        ws_epic_summary.cell(row, 2).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        ws_epic_summary.cell(row, 3).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        row += 1

        for epic in all_epics:
            epic_metrics_for_epic = [m['lead_time'] for m in epic_metrics if m['epic'] == epic]
            if epic_metrics_for_epic:
                ws_epic_summary.cell(row, 1).value = epic
                ws_epic_summary.cell(row, 2).value = round(mean(epic_metrics_for_epic), 2)
                ws_epic_summary.cell(row, 3).value = len(set(m['sprint'] for m in epic_metrics if m['epic'] == epic))
                ws_epic_summary.cell(row, 2).number_format = '0.00'
                ws_epic_summary.cell(row, 1).border = bord
                ws_epic_summary.cell(row, 2).border = bord
                ws_epic_summary.cell(row, 3).border = bord
                row += 1

        ws_epic_summary.column_dimensions['A'].width = 28
        ws_epic_summary.column_dimensions['B'].width = 20
        ws_epic_summary.column_dimensions['C'].width = 12

    # ========== Last Sprint Detail ==========
    if last_sprint:
        ws_last = wb_out.create_sheet(f"11. Detalle {last_sprint}")

        last_sprint_tickets = sorted([t for t in data if t['sprint'] == last_sprint], key=lambda x: x['key'])

        ws_last.cell(1, 1).value = f"Detalle: {last_sprint} ({len(last_sprint_tickets)} tickets)"
        ws_last.cell(1, 1).font = Font(bold=True, size=12)

        # Headers
        headers = ['Key', 'Pts', 'Lead Time', 'Dev', 'Code Review', 'Merged', 'Ready QA', 'QA', 'Sub-bugs']
        for col, header in enumerate(headers, 1):
            cell = ws_last.cell(3, col)
            cell.value = header
            cell.fill = h_fill
            cell.font = h_font
            cell.border = bord
            cell.alignment = cent

        # Detail rows
        r = 4
        for t in last_sprint_tickets:
            ws_last.cell(r, 1).value = t['key']
            ws_last.cell(r, 2).value = t['pts']
            ws_last.cell(r, 3).value = t['lead_time']
            ws_last.cell(r, 4).value = t['dev']
            ws_last.cell(r, 5).value = t['cr']
            ws_last.cell(r, 6).value = t['merged']
            ws_last.cell(r, 7).value = t['rqa']
            ws_last.cell(r, 8).value = t['qa']
            ws_last.cell(r, 9).value = t['bugs']

            for col in range(1, 10):
                ws_last.cell(r, col).border = bord
                if col > 1:
                    ws_last.cell(r, col).alignment = cent
                    if col > 2:
                        ws_last.cell(r, col).number_format = '0.00'
            r += 1

        # Summary stats
        summary_row = r + 2
        ws_last.cell(summary_row, 1).value = "RESUMEN"
        ws_last.cell(summary_row, 1).font = Font(bold=True, size=11)
        ws_last.cell(summary_row, 1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        summary_row += 1

        def safe_mean(values):
            valid = [v for v in values if v is not None]
            return round(mean(valid), 2) if valid else 0

        metrics_summary = [
            ("# Tickets", len(last_sprint_tickets)),
            ("Lead Time Promedio (días)", safe_mean([t['lead_time'] for t in last_sprint_tickets])),
            ("Tiempo Dev Promedio", safe_mean([t['dev'] for t in last_sprint_tickets])),
            ("Tiempo Code Review Promedio", safe_mean([t['cr'] for t in last_sprint_tickets])),
            ("Tiempo Merged Promedio", safe_mean([t['merged'] for t in last_sprint_tickets])),
            ("Tiempo Ready QA Promedio", safe_mean([t['rqa'] for t in last_sprint_tickets])),
            ("Tiempo QA Promedio", safe_mean([t['qa'] for t in last_sprint_tickets])),
            ("Sub-bugs Total", sum([t['bugs'] for t in last_sprint_tickets])),
            ("Sub-bugs Promedio/Ticket", safe_mean([t['bugs'] for t in last_sprint_tickets])),
        ]

        for label, value in metrics_summary:
            ws_last.cell(summary_row, 1).value = label
            ws_last.cell(summary_row, 2).value = value
            ws_last.cell(summary_row, 1).border = bord
            ws_last.cell(summary_row, 2).border = bord
            if isinstance(value, float):
                ws_last.cell(summary_row, 2).number_format = '0.00'
            summary_row += 1

        # Estado distribution
        summary_row += 1
        ws_last.cell(summary_row, 1).value = "DISTRIBUCIÓN POR CAMINO EN QA"
        ws_last.cell(summary_row, 1).font = Font(bold=True, size=11)
        ws_last.cell(summary_row, 1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        summary_row += 1

        with_in_qa = len([t for t in last_sprint_tickets if t['has_in_qa']])
        with_ready_qa = len([t for t in last_sprint_tickets if t['has_ready_qa'] and not t['has_in_qa']])
        without_qa = len([t for t in last_sprint_tickets if not t['has_ready_qa'] and not t['has_in_qa']])

        total = len(last_sprint_tickets)

        state_data = [
            ("Pasó por In QA (Ready QA + In QA)", with_in_qa, round(100 * with_in_qa / total, 1) if total > 0 else 0),
            ("Pasó por Ready QA (sin In QA)", with_ready_qa, round(100 * with_ready_qa / total, 1) if total > 0 else 0),
            ("Sin QA (saltó directo a Done)", without_qa, round(100 * without_qa / total, 1) if total > 0 else 0),
        ]

        chart_row = summary_row
        for label, count, pct in state_data:
            ws_last.cell(summary_row, 1).value = label
            ws_last.cell(summary_row, 2).value = f"{count} ({pct}%)"
            ws_last.cell(summary_row, 1).border = bord
            ws_last.cell(summary_row, 2).border = bord
            summary_row += 1

        # Pie chart
        pie = PieChart()
        pie.title = f"Distribución de Caminos en QA"
        pie_data = Reference(ws_last, min_col=2, min_row=chart_row, max_row=summary_row-1)
        pie_labels = Reference(ws_last, min_col=1, min_row=chart_row, max_row=summary_row-1)
        pie.add_data(pie_data)
        pie.set_categories(pie_labels)
        pie.height = 10
        pie.width = 14
        ws_last.add_chart(pie, "D3")

        for col in range(1, 3):
            ws_last.column_dimensions[get_column_letter(col)].width = 28

    return wb_out


# ══════════════════════════════════════════════════════════════════════════════
# ADVANCED METRICS REPORT (Velocity, Bottlenecks, Flow, Quality)
# ══════════════════════════════════════════════════════════════════════════════

def build_advanced_metrics_report(wb_out, data):
    """Add advanced metrics sheets: velocity, bottlenecks, flow, quality"""
    log.info("=== Adding Advanced Metrics ===")

    h_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    h_font = Font(bold=True, color="FFFFFF")
    bord = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    cent = Alignment(horizontal='center')

    # Group by sprint
    sprints_data = defaultdict(list)
    for t in data:
        sprints_data[t['sprint']].append(t)

    all_sprints = sorted(sprints_data.keys(), key=lambda x: str(x).lower())

    # ========== DASHBOARD ==========
    ws_dash = wb_out.create_sheet("0. Dashboard", 0)

    ws_dash.cell(1, 1).value = "DASHBOARD - VELOCIDAD Y PRODUCTIVIDAD"
    ws_dash.cell(1, 1).font = Font(bold=True, size=14)
    ws_dash.merge_cells('A1:D1')

    row = 3

    # Velocity (story points) per sprint
    ws_dash.cell(row, 1).value = "VELOCITY (Story Points por Sprint)"
    ws_dash.cell(row, 1).font = Font(bold=True, size=11)
    row += 1

    ws_dash.cell(row, 1).value = "Sprint"
    ws_dash.cell(row, 2).value = "Story Points"
    ws_dash.cell(row, 3).value = "# Tickets"
    ws_dash.cell(row, 4).value = "Avg SP/Ticket"
    for c in range(1, 5):
        ws_dash.cell(row, c).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        ws_dash.cell(row, c).border = bord
    row += 1

    velocity_data = []
    for sprint in all_sprints:
        tickets = sprints_data[sprint]
        total_sp = sum(t['pts'] for t in tickets if t['pts'])
        num_tickets = len(tickets)
        avg_sp = total_sp / num_tickets if num_tickets > 0 else 0

        ws_dash.cell(row, 1).value = sprint
        ws_dash.cell(row, 2).value = total_sp
        ws_dash.cell(row, 3).value = num_tickets
        ws_dash.cell(row, 4).value = round(avg_sp, 2)

        for c in range(1, 5):
            ws_dash.cell(row, c).border = bord
            if c > 1:
                ws_dash.cell(row, c).alignment = cent
                if c == 4:
                    ws_dash.cell(row, c).number_format = '0.00'

        velocity_data.append({'sprint': sprint, 'sp': total_sp, 'count': num_tickets})
        row += 1

    # Velocity chart
    chart_row = row
    chart = BarChart()
    chart.title = "Story Points Completados por Sprint"
    chart.y_axis.title = 'Story Points'
    data_ref = Reference(ws_dash, min_col=2, min_row=chart_row-len(velocity_data)-1, max_row=chart_row-1)
    cat_ref = Reference(ws_dash, min_col=1, min_row=chart_row-len(velocity_data), max_row=chart_row-1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cat_ref)
    chart.height = 12
    chart.width = 18
    ws_dash.add_chart(chart, "F3")

    # Throughput and completion
    row += 2
    ws_dash.cell(row, 1).value = "THROUGHPUT Y COMPLETITUD"
    ws_dash.cell(row, 1).font = Font(bold=True, size=11)
    row += 1

    ws_dash.cell(row, 1).value = "Métrica"
    ws_dash.cell(row, 2).value = "Valor"
    for c in range(1, 3):
        ws_dash.cell(row, c).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    row += 1

    total_sp_all = sum(t['sp'] for t in velocity_data)
    total_tickets_all = sum(t['count'] for t in velocity_data)
    avg_sp_all = total_sp_all / total_tickets_all if total_tickets_all > 0 else 0

    metrics_summary = [
        ("Tickets completados (total)", total_tickets_all),
        ("Story Points completados (total)", total_sp_all),
        ("SP promedio por ticket", round(avg_sp_all, 2)),
        ("SP promedio por sprint", round(total_sp_all / len(velocity_data), 2) if velocity_data else 0),
        ("Tickets promedio por sprint", round(total_tickets_all / len(velocity_data), 2) if velocity_data else 0),
    ]

    for label, val in metrics_summary:
        ws_dash.cell(row, 1).value = label
        ws_dash.cell(row, 2).value = val
        ws_dash.cell(row, 1).border = bord
        ws_dash.cell(row, 2).border = bord
        if isinstance(val, float):
            ws_dash.cell(row, 2).number_format = '0.00'
        row += 1

    ws_dash.column_dimensions['A'].width = 32
    ws_dash.column_dimensions['B'].width = 18

    # ========== BOTTLENECK ANALYSIS ==========
    ws_bn = wb_out.create_sheet("12. Cuellos de Botella")

    ws_bn.cell(1, 1).value = "ANÁLISIS DE CUELLOS DE BOTELLA"
    ws_bn.cell(1, 1).font = Font(bold=True, size=14)
    ws_bn.merge_cells('A1:D1')

    row = 3
    ws_bn.cell(row, 1).value = "Estado"
    ws_bn.cell(row, 2).value = "Tiempo Promedio (días)"
    ws_bn.cell(row, 3).value = "Máximo (días)"
    ws_bn.cell(row, 4).value = "Mínimo (días)"
    for c in range(1, 5):
        ws_bn.cell(row, c).fill = h_fill
        ws_bn.cell(row, c).font = h_font
        ws_bn.cell(row, c).border = bord
    row += 1

    state_times = defaultdict(list)
    for t in data:
        if t['dev']:
            state_times['Development'].append(t['dev'])
        if t['cr']:
            state_times['Code Review'].append(t['cr'])
        if t['merged']:
            state_times['Merged'].append(t['merged'])
        if t['rqa']:
            state_times['Ready QA'].append(t['rqa'])
        if t['qa']:
            state_times['QA'].append(t['qa'])

    bottleneck_list = []
    for state in ['Development', 'Code Review', 'Merged', 'Ready QA', 'QA']:
        if state in state_times:
            times = state_times[state]
            avg_time = mean(times)
            max_time = max(times)
            min_time = min(times)

            ws_bn.cell(row, 1).value = state
            ws_bn.cell(row, 2).value = round(avg_time, 2)
            ws_bn.cell(row, 3).value = round(max_time, 2)
            ws_bn.cell(row, 4).value = round(min_time, 2)

            for c in range(1, 5):
                ws_bn.cell(row, c).border = bord
                if c > 1:
                    ws_bn.cell(row, c).alignment = cent
                    ws_bn.cell(row, c).number_format = '0.00'

            bottleneck_list.append({'state': state, 'avg': avg_time})
            row += 1

    # Bottleneck chart
    chart_row = row
    chart = BarChart()
    chart.title = "Tiempo Promedio por Estado (Cuello de Botella)"
    chart.y_axis.title = 'Días'
    data_ref = Reference(ws_bn, min_col=2, min_row=3, max_row=row-1)
    cat_ref = Reference(ws_bn, min_col=1, min_row=4, max_row=row-1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cat_ref)
    chart.height = 12
    chart.width = 18
    ws_bn.add_chart(chart, "F3")

    # Time distribution pie
    row += 2
    ws_bn.cell(row, 1).value = "DISTRIBUCIÓN DE TIEMPO POR ESTADO"
    ws_bn.cell(row, 1).font = Font(bold=True, size=11)
    row += 1

    ws_bn.cell(row, 1).value = "Estado"
    ws_bn.cell(row, 2).value = "% Tiempo Total"
    for c in range(1, 3):
        ws_bn.cell(row, c).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    row += 1

    time_dist_row = row
    total_time = sum(mean(times) * len(times) for times in state_times.values() if times)
    for state in ['Development', 'Code Review', 'Merged', 'Ready QA', 'QA']:
        if state in state_times and state_times[state]:
            total_state_time = mean(state_times[state]) * len(state_times[state])
            pct = (total_state_time / total_time * 100) if total_time > 0 else 0

            ws_bn.cell(row, 1).value = state
            ws_bn.cell(row, 2).value = round(pct, 1)
            ws_bn.cell(row, 1).border = bord
            ws_bn.cell(row, 2).border = bord
            ws_bn.cell(row, 2).number_format = '0.0'
            row += 1

    pie = PieChart()
    pie.title = "% Tiempo Invertido por Estado"
    pie_data = Reference(ws_bn, min_col=2, min_row=time_dist_row-1, max_row=row-1)
    pie_labels = Reference(ws_bn, min_col=1, min_row=time_dist_row, max_row=row-1)
    pie.add_data(pie_data)
    pie.set_categories(pie_labels)
    pie.height = 10
    pie.width = 14
    ws_bn.add_chart(pie, "F" + str(time_dist_row))

    ws_bn.column_dimensions['A'].width = 20
    for c in range(2, 5):
        ws_bn.column_dimensions[get_column_letter(c)].width = 18

    # ========== VARIABILITY ANALYSIS ==========
    ws_var = wb_out.create_sheet("13. Variabilidad Lead Time")

    ws_var.cell(1, 1).value = "ANÁLISIS DE VARIABILIDAD Y PREDICTIBILIDAD"
    ws_var.cell(1, 1).font = Font(bold=True, size=14)
    ws_var.merge_cells('A1:F1')

    row = 3

    # Overall lead time stats
    ws_var.cell(row, 1).value = "Lead Time General"
    ws_var.cell(row, 1).font = Font(bold=True)
    row += 1

    ws_var.cell(row, 1).value = "Métrica"
    ws_var.cell(row, 2).value = "Valor (días)"
    for c in range(1, 3):
        ws_var.cell(row, c).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    row += 1

    lead_times = [t['lead_time'] for t in data if t['lead_time']]
    if lead_times:
        lead_times_sorted = sorted(lead_times)
        p95_idx = int(len(lead_times_sorted) * 0.95)
        p50_idx = int(len(lead_times_sorted) * 0.50)

        stats = [
            ("Promedio", mean(lead_times)),
            ("Mínimo", min(lead_times)),
            ("Máximo", max(lead_times)),
            ("Mediana (P50)", lead_times_sorted[p50_idx]),
            ("Percentil 95", lead_times_sorted[p95_idx]),
            ("Desv. Estándar", __import__('statistics').stdev(lead_times) if len(lead_times) > 1 else 0),
        ]

        for label, val in stats:
            ws_var.cell(row, 1).value = label
            ws_var.cell(row, 2).value = round(val, 2)
            ws_var.cell(row, 1).border = bord
            ws_var.cell(row, 2).border = bord
            ws_var.cell(row, 2).number_format = '0.00'
            row += 1

    # Lead time by story points
    row += 1
    ws_var.cell(row, 1).value = "Lead Time por Story Points"
    ws_var.cell(row, 1).font = Font(bold=True)
    row += 1

    ws_var.cell(row, 1).value = "Story Points"
    ws_var.cell(row, 2).value = "Promedio"
    ws_var.cell(row, 3).value = "Mín"
    ws_var.cell(row, 4).value = "Máx"
    ws_var.cell(row, 5).value = "P95"
    ws_var.cell(row, 6).value = "# Tickets"
    for c in range(1, 7):
        ws_var.cell(row, c).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    row += 1

    points_data = defaultdict(list)
    for t in data:
        if t['lead_time']:
            points_data[t['pts']].append(t['lead_time'])

    for pts in sorted(points_data.keys()):
        times = points_data[pts]
        times_sorted = sorted(times)
        p95_idx = int(len(times_sorted) * 0.95)

        ws_var.cell(row, 1).value = pts
        ws_var.cell(row, 2).value = round(mean(times), 2)
        ws_var.cell(row, 3).value = round(min(times), 2)
        ws_var.cell(row, 4).value = round(max(times), 2)
        ws_var.cell(row, 5).value = round(times_sorted[p95_idx], 2)
        ws_var.cell(row, 6).value = len(times)

        for c in range(1, 7):
            ws_var.cell(row, c).border = bord
            if c > 1:
                ws_var.cell(row, c).alignment = cent
                if c != 6:
                    ws_var.cell(row, c).number_format = '0.00'
        row += 1

    ws_var.column_dimensions['A'].width = 20
    for c in range(2, 7):
        ws_var.column_dimensions[get_column_letter(c)].width = 14

    # ========== FLOW & WIP ANALYSIS ==========
    ws_flow = wb_out.create_sheet("14. Flujo y Work-in-Progress")

    ws_flow.cell(1, 1).value = "ANÁLISIS DE FLUJO DE TRABAJO"
    ws_flow.cell(1, 1).font = Font(bold=True, size=14)
    ws_flow.merge_cells('A1:C1')

    row = 3
    ws_flow.cell(row, 1).value = "Estado"
    ws_flow.cell(row, 2).value = "# Tickets Activos"
    ws_flow.cell(row, 3).value = "% del Total"
    for c in range(1, 4):
        ws_flow.cell(row, c).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    row += 1

    # Count tickets by state (using states_visited)
    state_counts = defaultdict(int)
    for t in data:
        if 'In Progress' in t['states_visited']:
            state_counts['In Progress'] += 1
        if 'Code Review' in t['states_visited']:
            state_counts['Code Review'] += 1
        if 'Merged' in t['states_visited']:
            state_counts['Merged'] += 1
        if 'Ready for QA' in t['states_visited']:
            state_counts['Ready for QA'] += 1
        if 'In QA' in t['states_visited']:
            state_counts['In QA'] += 1

    total_active = len(data)
    flow_row = row

    for state in ['In Progress', 'Code Review', 'Merged', 'Ready for QA', 'In QA']:
        count = state_counts[state]
        pct = (count / total_active * 100) if total_active > 0 else 0

        ws_flow.cell(row, 1).value = state
        ws_flow.cell(row, 2).value = count
        ws_flow.cell(row, 3).value = round(pct, 1)

        for c in range(1, 4):
            ws_flow.cell(row, c).border = bord
            if c > 1:
                ws_flow.cell(row, c).alignment = cent
                if c == 3:
                    ws_flow.cell(row, c).number_format = '0.0'
        row += 1

    # Flow chart
    chart = BarChart()
    chart.title = "Tickets en Cada Estado del Pipeline"
    chart.y_axis.title = '# Tickets'
    data_ref = Reference(ws_flow, min_col=2, min_row=flow_row-1, max_row=row-1)
    cat_ref = Reference(ws_flow, min_col=1, min_row=flow_row, max_row=row-1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cat_ref)
    chart.height = 10
    chart.width = 16
    ws_flow.add_chart(chart, "E3")

    # Throughput per week or sprint
    row += 2
    ws_flow.cell(row, 1).value = "THROUGHPUT POR SPRINT"
    ws_flow.cell(row, 1).font = Font(bold=True)
    row += 1

    ws_flow.cell(row, 1).value = "Sprint"
    ws_flow.cell(row, 2).value = "# Tickets Cerrados"
    ws_flow.cell(row, 3).value = "Story Points"
    for c in range(1, 4):
        ws_flow.cell(row, c).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    row += 1

    throughput_row = row
    for sprint in all_sprints:
        tickets = sprints_data[sprint]
        total_sp = sum(t['pts'] for t in tickets if t['pts'])

        ws_flow.cell(row, 1).value = sprint
        ws_flow.cell(row, 2).value = len(tickets)
        ws_flow.cell(row, 3).value = total_sp

        for c in range(1, 4):
            ws_flow.cell(row, c).border = bord
            if c > 1:
                ws_flow.cell(row, c).alignment = cent
        row += 1

    ws_flow.column_dimensions['A'].width = 20
    ws_flow.column_dimensions['B'].width = 18
    ws_flow.column_dimensions['C'].width = 14

    # ========== QUALITY ANALYSIS ==========
    ws_qual = wb_out.create_sheet("15. Análisis de Calidad")

    ws_qual.cell(1, 1).value = "ANÁLISIS DE CALIDAD Y CONFIABILIDAD"
    ws_qual.cell(1, 1).font = Font(bold=True, size=14)
    ws_qual.merge_cells('A1:D1')

    row = 3

    # Tickets with issues
    reopened_count = 0
    for t in data:
        if 'Bug Fixing' in t['states_visited']:
            reopened_count += 1

    ws_qual.cell(row, 1).value = "Métrica de Calidad"
    ws_qual.cell(row, 2).value = "Valor"
    for c in range(1, 3):
        ws_qual.cell(row, c).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    row += 1

    total_tickets = len(data)
    reopening_pct = (reopened_count / total_tickets * 100) if total_tickets > 0 else 0

    quality_metrics = [
        ("Total de Tickets", total_tickets),
        ("Tickets con Bug Fixing", reopened_count),
        ("Reopening Rate", round(reopening_pct, 1)),
        ("Tasa de Éxito (sin Bug Fixing)", round(100 - reopening_pct, 1)),
    ]

    for label, val in quality_metrics:
        ws_qual.cell(row, 1).value = label
        ws_qual.cell(row, 2).value = val
        ws_qual.cell(row, 1).border = bord
        ws_qual.cell(row, 2).border = bord
        if isinstance(val, float):
            ws_qual.cell(row, 2).number_format = '0.0'
        row += 1

    # Reopening by sprint
    row += 1
    ws_qual.cell(row, 1).value = "REOPENING RATE POR SPRINT"
    ws_qual.cell(row, 1).font = Font(bold=True)
    row += 1

    ws_qual.cell(row, 1).value = "Sprint"
    ws_qual.cell(row, 2).value = "Con Bug Fixing"
    ws_qual.cell(row, 3).value = "Total"
    ws_qual.cell(row, 4).value = "Reopening Rate"
    for c in range(1, 5):
        ws_qual.cell(row, c).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    row += 1

    reopen_row = row
    for sprint in all_sprints:
        tickets = sprints_data[sprint]
        reopen_in_sprint = len([t for t in tickets if 'Bug Fixing' in t['states_visited']])
        total_in_sprint = len(tickets)
        reopen_rate = (reopen_in_sprint / total_in_sprint * 100) if total_in_sprint > 0 else 0

        ws_qual.cell(row, 1).value = sprint
        ws_qual.cell(row, 2).value = reopen_in_sprint
        ws_qual.cell(row, 3).value = total_in_sprint
        ws_qual.cell(row, 4).value = round(reopen_rate, 1)

        for c in range(1, 5):
            ws_qual.cell(row, c).border = bord
            if c > 1:
                ws_qual.cell(row, c).alignment = cent
                if c == 4:
                    ws_qual.cell(row, c).number_format = '0.0'
        row += 1

    ws_qual.column_dimensions['A'].width = 24
    for c in range(2, 5):
        ws_qual.column_dimensions[get_column_letter(c)].width = 16

    log.info("  Advanced metrics sheets created successfully")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    t0 = time.time()
    log.info("============================================")
    log.info("  Jira Ticket Report Generator v4 (GitHub)  ")
    log.info("============================================")

    if not JIRA_EMAIL or not JIRA_API_TOKEN:
        log.error("JIRA_EMAIL and JIRA_API_TOKEN environment variables must be set!")
        return

    # Fetch and process
    issues = fetch_all_issues()
    if not issues:
        log.error("No issues found. Check credentials and JQL.")
        return

    rows   = []
    failed = 0
    total  = len(issues)
    done   = 0

    log.info("=== Processing %d tickets with %d threads ===", total, MAX_WORKERS)

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {pool.submit(process_issue, issue): issue.get("key", "?")
                   for issue in issues}

        for future in as_completed(futures):
            done += 1
            result = future.result()
            if result:
                rows.append(result)
            else:
                failed += 1
            if done % 25 == 0 or done == total:
                log.info("  %d/%d (%.0f%%)  rows=%d  failed=%d",
                         done, total, done / total * 100, len(rows), failed)

    log.info("Done: %d rows, %d failed/skipped", len(rows), failed)

    # Write data report
    log.info("=== Writing Data Report Excel file... ===")
    wb_data = build_data_workbook(rows)
    wb_data.save(DATA_REPORT_PATH)

    # Write metrics report
    log.info("=== Writing Metrics Report Excel file... ===")
    wb_metrics = build_metrics_report_from_data(rows)

    # Parse data for advanced metrics (same processing as metrics report)
    data = []
    for row in rows:
        key = row[0]
        pts = row[2]
        sprint = row[13]
        num_subtasks = row[14] or 0
        num_subbugs = row[15] or 0
        parent = row[10]

        try:
            changelog_json = row[8]
            changelog = json.loads(changelog_json) if changelog_json else []
        except:
            changelog = []

        first_status_time = {}
        all_states = set()

        for change in changelog:
            created_str = change.get('created', '')
            for item in change.get('items', []):
                if item.get('field') == 'status':
                    to_status_str = item.get('toString')
                    all_states.add(to_status_str)
                    if to_status_str and to_status_str not in first_status_time:
                        try:
                            fixed_created = created_str[:-2] + ':' + created_str[-2:]
                            dt = datetime.fromisoformat(fixed_created)
                            first_status_time[to_status_str] = dt
                        except:
                            pass

        if 'In Progress' not in first_status_time or 'Done' not in first_status_time:
            continue
        if not pts:
            continue

        try:
            pts_int = int(pts)
        except:
            continue

        lead_time_days = (first_status_time['Done'] - first_status_time['In Progress']).total_seconds() / (24*3600)

        def time_between(from_st, to_st):
            if from_st in first_status_time and to_st in first_status_time:
                dt = (first_status_time[to_st] - first_status_time[from_st]).total_seconds() / (24*3600)
                return round(dt, 2)
            return None

        sprints = [s.strip() for s in str(sprint).split(',')] if sprint else []
        last_sprint = sprints[-1] if sprints else sprint

        total_bugs = num_subbugs + num_subtasks

        ticket = {
            'key': key,
            'sprint': last_sprint,
            'pts': pts_int,
            'lead_time': round(lead_time_days, 2),
            'dev': time_between('In Progress', 'Code Review'),
            'cr': time_between('Code Review', 'Merged'),
            'merged': time_between('Merged', 'Ready for QA'),
            'rqa': time_between('Ready for QA', 'In QA'),
            'qa': time_between('In QA', 'Done'),
            'bugs': total_bugs,
            'states_visited': all_states,
        }
        data.append(ticket)

    # Add advanced metrics
    log.info("=== Adding Advanced Metrics ===")
    build_advanced_metrics_report(wb_metrics, data)

    wb_metrics.save(METRICS_REPORT_PATH)

    elapsed = time.time() - t0
    log.info("=== Finished in %.1fs ===", elapsed)
    print(f"\n  Data Report saved to: {DATA_REPORT_PATH}")
    print(f"  Metrics Report saved to: {METRICS_REPORT_PATH}")
    print(f"    {len(rows)} tickets  |  {elapsed:.0f}s total\n")

if __name__ == "__main__":
    main()
