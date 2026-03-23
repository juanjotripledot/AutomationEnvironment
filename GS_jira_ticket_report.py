#!/usr/bin/env python3
"""
Jira Ticket Report Generator — v3 (GitHub Actions Edition)
────────────────────────────────────────────────────────────────────────────────
Changes vs Original:
  • Reads Jira credentials from Environment Variables (Secrets).
  • Outputs the Excel file to the local working directory (for GitHub Artifacts).
  • Removes local macOS Google Drive copying logic (incompatible with headless runners).
────────────────────────────────────────────────────────────────────────────────
"""

import json
import logging
import os
import time
from datetime import datetime, timedelta, timezone
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter



# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION — Read from GitHub Secrets (Environment Variables)
# ══════════════════════════════════════════════════════════════════════════════
JIRA_BASE_URL  = os.environ.get("JIRA_BASE_URL", "https://tripledotstudios.atlassian.net")
JIRA_EMAIL     = os.environ.get("JIRA_EMAIL")
JIRA_API_TOKEN = os.environ.get("JIRA_API_TOKEN")

# Save directly to the current working directory in GitHub Actions
OUTPUT_PATH = "GS_jira_year_report.xlsx"

# ══════════════════════════════════════════════════════════════════════════════
# TUNABLES
# ══════════════════════════════════════════════════════════════════════════════
MAX_WORKERS         = 8    # parallel threads for per-ticket requests
SEARCH_PAGE_SIZE    = 100  # tickets per search page (Jira max = 100)
CHANGELOG_PAGE_SIZE = 100  # changelog entries per page (Jira max = 100)
MAX_RETRIES         = 5
TIMEOUT             = 45   # seconds

# ══════════════════════════════════════════════════════════════════════════════
# DOMAIN CONSTANTS
# ══════════════════════════════════════════════════════════════════════════════
STATUSES = [
    "In Refinement",
    "Analyisis",                      # matches the Jira board typo
    "Need Info",
    "Re-Opened",
    "Ready for dev",
    "In Progress",
    "Changes Requested by reviewer",
    "Code Review",
    "In QA feature branch",
    "Verified on feature branch",
    "Staging",
    "In QA",
    "Verified for production",
]

TIME_IN_DEV_STATUSES = {
    "In Progress",
    "Re-Opened",
    "Code Review",
    "Deployed on Feature branch",
    "In QA feature branch",
    "Verified on feature branch",
}

# Column headers — order must match row assembly in process_issue()
HEADERS = [
    "Key",
    "Summary",
    "Story points",
    "Components",
    "Labels",
    "Assignee",
    "Description",
    "Comments",
    "Changelog (JSON)",
    "Linked Items",
    "Parent",
    "Depends on",
    "Time in Refinement",
    "Time in Analysis",
    "Time in Need Info",
    "Time in Re-Opened",
    "Time in Ready for dev",
    "Time in In Progress",
    "Time in Changes Requested by reviewer",
    "Time in Code Review",
    "Time in In QA feature branch",
    "Time in Verified on feature branch",
    "Time in Staging",
    "Time in In QA",
    "Time in Verified for production",
    "Time in dev",
    "Delivery Time",
    "Number of Re-opens",
]

# Maps the "Time in X" header text to exact Jira status name
STATUS_HEADER_MAP = {
    "Time in Refinement":                    "In Refinement",
    "Time in Analysis":                      "Analyisis",
    "Time in Need Info":                     "Need Info",
    "Time in Re-Opened":                     "Re-Opened",
    "Time in Ready for dev":                 "Ready for dev",
    "Time in In Progress":                   "In Progress",
    "Time in Changes Requested by reviewer": "Changes Requested by reviewer",
    "Time in Code Review":                   "Code Review",
    "Time in In QA feature branch":          "In QA feature branch",
    "Time in Verified on feature branch":    "Verified on feature branch",
    "Time in Staging":                       "Staging",
    "Time in In QA":                         "In QA",
    "Time in Verified for production":       "Verified for production",
}

# 0-based index of "Time in Refinement" in HEADERS
FIRST_STATUS_COL = HEADERS.index("Time in Refinement")

# ══════════════════════════════════════════════════════════════════════════════
# LOGGING
# ══════════════════════════════════════════════════════════════════════════════
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ══════════════════════════════════════════════════════════════════════════════
# HTTP SESSION
# ══════════════════════════════════════════════════════════════════════════════
SESSION = requests.Session()
if JIRA_EMAIL and JIRA_API_TOKEN:
    SESSION.auth = (JIRA_EMAIL, JIRA_API_TOKEN)
SESSION.headers.update({
    "Accept": "application/json",
    "Content-Type": "application/json",
})


def _get(url, params=None, attempt=0):
    """
    GET with exponential-backoff retry.
    Honours Retry-After on 429 responses.
    Returns parsed JSON or None on permanent failure.
    """
    try:
        resp = SESSION.get(url, params=params, timeout=TIMEOUT)

        if resp.status_code == 429 or resp.status_code >= 500:
            if attempt < MAX_RETRIES:
                wait = max(int(resp.headers.get("Retry-After", 0)), 2 ** attempt)
                log.warning("HTTP %s — backing off %ds (attempt %d/%d)",
                            resp.status_code, wait, attempt + 1, MAX_RETRIES)
                time.sleep(wait)
                return _get(url, params, attempt + 1)

        resp.raise_for_status()
        return resp.json()

    except requests.RequestException as exc:
        if attempt < MAX_RETRIES:
            wait = 2 ** attempt
            log.warning("Network error (%s) — retrying in %ds (attempt %d/%d)",
                        exc, wait, attempt + 1, MAX_RETRIES)
            time.sleep(wait)
            return _get(url, params, attempt + 1)

        log.error("Permanent failure for %s: %s", url, exc)
        return None


# ══════════════════════════════════════════════════════════════════════════════
# WORKING-HOURS CALCULATOR  (Mon-Fri, whole calendar day; no shift window)
# ══════════════════════════════════════════════════════════════════════════════

def working_hours_between(start, end):
    """
    Return working hours (Mon-Fri) between two timezone-aware datetimes.
    Saturday and Sunday contribute 0 hours.
    All hours within a weekday are counted (no 9-5 window assumed).
    """
    if end <= start:
        return 0.0

    total_secs = 0.0
    day = start.replace(hour=0, minute=0, second=0, microsecond=0)

    while day < end:
        next_day = day + timedelta(days=1)
        if day.weekday() < 5:          # 0=Mon, 1=Tue, ..., 4=Fri
            seg_start = max(start, day)
            seg_end   = min(end, next_day)
            if seg_end > seg_start:
                total_secs += (seg_end - seg_start).total_seconds()
        day = next_day

    return total_secs / 3600.0


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 - FETCH ALL ISSUES  (/rest/api/3/search/jql with cursor pagination)
# ══════════════════════════════════════════════════════════════════════════════

def fetch_all_issues():
    """
    Page through every issue matching the JQL using /rest/api/3/search/jql.
    """
    jql = (
        'project in ("GS","SSO","Backend Services") AND (LABELS NOT IN (NoKPI) OR LABELS = EMPTY) AND issuetype not in (Documentation,Epic) '
        'AND updated >= -365d '
        'ORDER BY updated DESC'
    )

    url             = f"{JIRA_BASE_URL}/rest/api/3/search/jql"
    issues          = []
    next_page_token = None
    offset          = 0
    page_num        = 0

    log.info("=== STEP 1: Fetching all issues via /rest/api/3/search/jql ===")

    while True:
        params = {
            "jql":        jql,
            "maxResults": SEARCH_PAGE_SIZE,
            "fields":     "key",
        }

        if next_page_token:
            params["nextPageToken"] = next_page_token
        else:
            params["startAt"] = offset

        data = _get(url, params=params)

        if not data:
            log.error("Empty response on page %d — stopping.", page_num)
            break

        batch = data.get("issues", [])
        if not batch:
            log.info("  Empty batch on page %d — reached the end.", page_num)
            break

        issues.extend(batch)
        page_num += 1

        total          = data.get("total", 0)
        next_page_token = data.get("nextPageToken") or data.get("next_page_token")

        log.info("  page=%-3d  batch=%-4d  accumulated=%-5d  total_reported=%-5s  cursor=%s",
                 page_num, len(batch), len(issues),
                 str(total) if total else "?",
                 "yes" if next_page_token else "no")

        if next_page_token:
            continue

        if total and len(issues) >= total:
            break
        if len(batch) < SEARCH_PAGE_SIZE:
            break

        offset += len(batch)

    log.info("Total issues collected: %d", len(issues))
    return issues


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 - FETCH FULL CHANGELOG  (/rest/api/3/issue/{key}/changelog)
# ══════════════════════════════════════════════════════════════════════════════

def fetch_changelog(key):
    url       = f"{JIRA_BASE_URL}/rest/api/3/issue/{key}/changelog"
    histories = []
    start     = 0

    while True:
        data = _get(url, params={"startAt": start, "maxResults": CHANGELOG_PAGE_SIZE})

        if not data:
            break

        batch = data.get("values", [])
        if not batch:
            break

        histories.extend(batch)
        total = data.get("total", 0)

        if len(histories) >= total:
            break

        start += len(batch)

    return histories


# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 - FETCH ISSUE FIELDS
# ══════════════════════════════════════════════════════════════════════════════

def fetch_issue_fields(key):
    url  = f"{JIRA_BASE_URL}/rest/api/3/issue/{key}"
    data = _get(url, params={
        "fields": "summary,customfield_10023,components,labels,assignee,description,comment,parent,issuelinks"
    })
    return (data or {}).get("fields", {})


# ══════════════════════════════════════════════════════════════════════════════
# ADF -> PLAIN TEXT
# ══════════════════════════════════════════════════════════════════════════════

def _adf_to_text(node):
    if not node:
        return ""
    if isinstance(node, str):
        return node
    if node.get("type") == "text":
        return node.get("text", "")
    parts = [_adf_to_text(c) for c in node.get("content", [])]
    return " ".join(p for p in parts if p)


# ══════════════════════════════════════════════════════════════════════════════
# COMMENTS -> JSON STRING
# ══════════════════════════════════════════════════════════════════════════════

def build_comments_json(fields):
    result = []
    try:
        for c in fields.get("comment", {}).get("comments", []):
            body_obj = c.get("body", {})
            text     = _adf_to_text(body_obj) if isinstance(body_obj, dict) else str(body_obj or "")
            result.append({
                "datetime": c.get("created", ""),
                "author":   (c.get("author") or {}).get("displayName", "Unknown"),
                "comment":  text,
            })
    except Exception as exc:
        log.warning("Comments parse error: %s", exc)
    return json.dumps(result, ensure_ascii=False)


# ══════════════════════════════════════════════════════════════════════════════
# CHANGELOG -> JSON STRING  (raw, for the Changelog column)
# ══════════════════════════════════════════════════════════════════════════════

def build_changelog_json(histories):
    result = []
    for h in histories:
        result.append({
            "id":      h.get("id", ""),
            "created": h.get("created", ""),
            "author":  (h.get("author") or {}).get("displayName", "Unknown"),
            "items": [
                {
                    "field":      item.get("field", ""),
                    "fieldtype":  item.get("fieldtype", ""),
                    "from":       item.get("from", ""),
                    "fromString": item.get("fromString", ""),
                    "to":         item.get("to", ""),
                    "toString":   item.get("toString", ""),
                }
                for item in h.get("items", [])
            ],
        })
    return json.dumps(result, ensure_ascii=False)


# ══════════════════════════════════════════════════════════════════════════════
# STATUS METRICS  (working hours)
# ══════════════════════════════════════════════════════════════════════════════

def _parse_ts(ts_str):
    if not ts_str:
        return None
    try:
        s = ts_str.strip()
        if s.endswith("Z"):
            s = s[:-1] + "+00:00"
        elif len(s) > 5 and s[-5] in "+-" and ":" not in s[-5:]:
            s = s[:-2] + ":" + s[-2:]
        return datetime.fromisoformat(s).astimezone(timezone.utc)
    except (ValueError, TypeError):
        return None


def compute_status_metrics(histories):
    now         = datetime.now(timezone.utc)
    transitions = []

    for h in histories:
        ts = _parse_ts(h.get("created", ""))
        if ts is None:
            continue
        for item in h.get("items", []):
            if item.get("field") == "status":
                transitions.append({
                    "ts":   ts,
                    "from": item.get("fromString", ""),
                    "to":   item.get("toString", ""),
                })

    transitions.sort(key=lambda x: x["ts"])

    time_in           = {s: 0.0 for s in STATUSES}
    first_in_progress = None
    first_verified    = None
    reopen_count      = 0

    for i, t in enumerate(transitions):
        name      = t["to"]
        enter     = t["ts"]
        exit_time = transitions[i + 1]["ts"] if i + 1 < len(transitions) else now

        wh = working_hours_between(enter, exit_time)

        if name in time_in:
            time_in[name] += wh

        if name == "In Progress" and first_in_progress is None:
            first_in_progress = enter
        if name == "Verified for production" and first_verified is None:
            first_verified = enter
        if name == "Re-Opened":
            reopen_count += 1

    time_in_dev = sum(time_in.get(s, 0.0) for s in TIME_IN_DEV_STATUSES)

    delivery_time = None
    if first_in_progress and first_verified and first_verified > first_in_progress:
        delivery_time = working_hours_between(first_in_progress, first_verified)

    return time_in, time_in_dev, delivery_time, reopen_count


# ══════════════════════════════════════════════════════════════════════════════
# PROCESS ONE ISSUE  (called from thread pool)
# ══════════════════════════════════════════════════════════════════════════════

def process_issue(basic_issue):
    key = basic_issue.get("key", "UNKNOWN")
    try:
        fields = fetch_issue_fields(key)
        if not fields:
            log.warning("No fields for %s — skipping", key)
            return None

        summary     = fields.get("summary", "")
        sp          = fields.get("customfield_10023")
        components  = ", ".join(c.get("name", "") for c in (fields.get("components") or []))
        labels      = ", ".join(fields.get("labels") or [])
        assignee    = (fields.get("assignee") or {}).get("displayName", "")
        desc_obj    = fields.get("description")
        description = _adf_to_text(desc_obj) if isinstance(desc_obj, dict) else (desc_obj or "")
        comments    = build_comments_json(fields)

        current_project_key = key.split("-")[0] if "-" in key else ""
        depends_on_projects = set()

        # Extract Parent
        parent_key = ""
        parent_field = fields.get("parent")
        if parent_field:
            parent_key = parent_field.get("key", "")
            if parent_key:
                parent_proj = parent_key.split("-")[0]
                if parent_proj and parent_proj != current_project_key:
                    depends_on_projects.add(parent_proj)

        # Extract Linked Items
        linked_items_list = []
        issue_links = fields.get("issuelinks", [])
        for link in issue_links:
            link_type_name = link.get("type", {}).get("name", "")
            if "inwardIssue" in link:
                linked_issue = link["inwardIssue"]
                direction = link.get("type", {}).get("inward", link_type_name)
            elif "outwardIssue" in link:
                linked_issue = link["outwardIssue"]
                direction = link.get("type", {}).get("outward", link_type_name)
            else:
                continue
            
            linked_key = linked_issue.get("key", "")
            linked_items_list.append({
                "relationship": direction,
                "key": linked_key,
                "summary": (linked_issue.get("fields") or {}).get("summary", ""),
                "status": ((linked_issue.get("fields") or {}).get("status") or {}).get("name", "")
            })
            if linked_key:
                linked_proj = linked_key.split("-")[0]
                if linked_proj and linked_proj != current_project_key:
                    depends_on_projects.add(linked_proj)
                    
        linked_items_json = json.dumps(linked_items_list, ensure_ascii=False)
        depends_on_str = ", ".join(sorted(list(depends_on_projects)))

        histories      = fetch_changelog(key)
        changelog_json = build_changelog_json(histories)

        time_in, time_in_dev, delivery_time, reopen_count = compute_status_metrics(histories)

        row = [
            key,
            summary,
            sp,
            components,
            labels,
            assignee,
            description,
            comments,
            changelog_json,
            linked_items_json,
            parent_key,
            depends_on_str,
        ]

        for header in HEADERS[FIRST_STATUS_COL: FIRST_STATUS_COL + len(STATUS_HEADER_MAP)]:
            val = time_in.get(STATUS_HEADER_MAP[header], 0.0)
            row.append(round(val, 2) if val else None)

        row.append(round(time_in_dev, 2)   if time_in_dev    else None)
        row.append(round(delivery_time, 2) if delivery_time  else None)
        row.append(reopen_count            if reopen_count   else None)

        return row

    except Exception as exc:
        log.error("Error processing %s: %s", key, exc, exc_info=True)
        return None


# ══════════════════════════════════════════════════════════════════════════════
# BUILD EXCEL WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════

import re

# Match illegal XML 1.0 control characters (allows tab, newline, carriage return)
ILLEGAL_XML_CHARS_RE = re.compile(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]")

def clean_for_excel(val):
    if isinstance(val, str):
        return ILLEGAL_XML_CHARS_RE.sub("", val)
    return val

def build_workbook(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Jira Tickets"

    h_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    h_fill  = PatternFill("solid", start_color="1F4E79")
    h_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.append(HEADERS)
    for col in range(1, len(HEADERS) + 1):
        cell           = ws.cell(row=1, column=col)
        cell.font      = h_font
        cell.fill      = h_fill
        cell.alignment = h_align
    ws.row_dimensions[1].height = 32

    d_font     = Font(name="Arial", size=10)
    d_align    = Alignment(wrap_text=True, vertical="top")
    num_format = "#,##0.00"
    numeric_start_col = FIRST_STATUS_COL + 1   # 1-based

    for row_data in rows:
        cleaned_row = [clean_for_excel(v) for v in row_data]
        ws.append(cleaned_row)
        r = ws.max_row
        for col in range(1, len(HEADERS) + 1):
            cell           = ws.cell(row=r, column=col)
            cell.font      = d_font
            cell.alignment = d_align
            if col >= numeric_start_col:
                cell.number_format = num_format

    widths = {
        1: 16,   # Key
        2: 42,   # Summary
        3: 12,   # Story points
        4: 22,   # Components
        5: 22,   # Labels
        6: 22,   # Assignee
        7: 55,   # Description
        8: 55,   # Comments
        9: 60,   # Changelog (JSON)
        10: 60,  # Linked Items
        11: 16,  # Parent
        12: 24,  # Depends on
    }
    for col in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(col, 20)

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = ws.dimensions

    return wb




# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    t0 = time.time()
    log.info("============================================")
    log.info("  Jira Ticket Report Generator (GitHub)     ")
    log.info("============================================")

    if not JIRA_EMAIL or not JIRA_API_TOKEN:
        log.error("JIRA_EMAIL and JIRA_API_TOKEN environment variables must be set!")
        return

    issues = fetch_all_issues()
    if not issues:
        log.error("No issues found. Check credentials and JQL.")
        return

    rows   = []
    failed = 0
    total  = len(issues)
    done   = 0

    log.info("=== Processing %d tickets with %d threads ===", total, MAX_WORKERS)

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {pool.submit(process_issue, issue): issue.get("key", "?")
                   for issue in issues}

        for future in as_completed(futures):
            done += 1
            result = future.result()
            if result:
                rows.append(result)
            else:
                failed += 1
            if done % 25 == 0 or done == total:
                log.info("  %d/%d (%.0f%%)  rows=%d  failed=%d",
                         done, total, done / total * 100, len(rows), failed)

    log.info("Done: %d rows, %d failed/skipped", len(rows), failed)

    log.info("=== Writing Excel file... ===")
    wb = build_workbook(rows)
    wb.save(OUTPUT_PATH)

    elapsed = time.time() - t0
    log.info("=== Finished in %.1fs ===", elapsed)
    print(f"\n  Report saved to: {OUTPUT_PATH}")
    print(f"    {len(rows)} tickets  |  {elapsed:.0f}s total\n")

if __name__ == "__main__":
    main()
