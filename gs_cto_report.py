"""
GS CTO Monthly Performance Report
=================================
Genera un PDF con 5 páginas replicando el informe mensual del CTO:

  Página 1 - Portada + Introducción
  Página 2 - Time in Development (histogramas 3/5/8/13 SP)
  Página 3 - Time in QA          (histogramas 3/5/8/13 SP)
  Página 4 - Team Velocity        (líneas bisemanal + mensual)
  Página 5 - Effort Distribution  (donut global + barras mensuales y trimestrales
                                    por estudio y por juego)

Definiciones:
  Time in Dev = Time in Re-Opened + Time in In Progress + Time in Code Review
  Time in QA  = Time in Staging + Time in In QA
  Los valores de la planilla están en HORAS laborables → se dividen entre 8.

Período de referencia para Dev/QA/Velocity:
  Fecha de "Verified for Production" del ticket.

Las páginas 4 y 5 toman las últimas 10 etiquetas bisemanales/mensuales y 8
trimestrales, ancladas al último mes completo (mes anterior a la ejecución).

Uso:
  python gs_cto_report.py \\
      --xlsx GS_jira_year_report.xlsx \\
      --mapping GameStudios.csv \\
      --output GS_CTO_Report.pdf \\
      [--month 2026-03]

Dependencias:
  pip install openpyxl matplotlib reportlab Pillow
"""

import argparse
import csv
import json
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import openpyxl
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    BaseDocTemplate, Frame, PageTemplate, Paragraph, Spacer,
    Image, Table, TableStyle, PageBreak
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT

# ══════════════════════════════════════════════════════════════════════════════
# Constantes
# ══════════════════════════════════════════════════════════════════════════════

SP_VALID = {3, 5, 8, 13}

# Status names (uppercased for comparisons)
STATUS_IN_PROGRESS = "IN PROGRESS"
STATUS_STAGING     = "STAGING"
STATUS_IN_QA       = "IN QA"

COL_KEY            = "KEY"
COL_STORY_POINTS   = "STORY POINTS"
COL_COMPONENTS     = "COMPONENTS"
COL_CHANGELOG_JSON = "CHANGELOG (JSON)"

# Prioridad para determinar la fecha de "finalización" del ticket (período)
# y también como límite derecho para el cálculo de Time in QA.
VERIFIED_STATUSES_PRIORITY = ["Verified for Production", "Production", "Done"]

PERIODS_BACK_BIWEEKLY = 10
PERIODS_BACK_MONTHLY  = 10
PERIODS_BACK_QUARTERLY = 8

OUTLIER_IQR_MULT = 2.5

# Paleta de colores
COLOR_BAR     = "#1e6fc9"
COLOR_OUTLIER = "#e84c30"
COLOR_LINE    = "#1e6fc9"

STUDIO_COLORS = [
    "#1e6fc9", "#2da06e", "#e8a020", "#e84c30", "#8b4fc9",
    "#26b3b3", "#c94f8b", "#6b8e23", "#a07050", "#888780",
]
GAME_COLORS = STUDIO_COLORS  # mismo set


# ══════════════════════════════════════════════════════════════════════════════
# Helpers de fechas
# ══════════════════════════════════════════════════════════════════════════════

def previous_month_bounds(today: date = None):
    """Devuelve (first_day, last_day, label) del mes anterior a today."""
    if today is None:
        today = date.today()
    first_of_this_month = date(today.year, today.month, 1)
    last_of_prev = first_of_this_month - timedelta(days=1)
    first_of_prev = date(last_of_prev.year, last_of_prev.month, 1)
    label = last_of_prev.strftime("%B %Y")
    return first_of_prev, last_of_prev, label


def parse_ts(ts_str):
    if not ts_str:
        return None
    s = str(ts_str).strip()
    if s.endswith("Z"):
        s = s[:-1] + "+00:00"
    try:
        return datetime.fromisoformat(s)
    except ValueError:
        return None


def working_days_between(start_dt: datetime, end_dt: datetime) -> float:
    """
    Devuelve los días laborables entre dos datetimes (solo L-V).
    Cuenta fracciones de día. Los fines de semana (sábado/domingo) se
    excluyen completamente.

    Ejemplo: viernes 17:00 → lunes 09:00 = 16h laborables
      (viernes 17→24 = 7h, sábado y domingo = 0h, lunes 0→9 = 9h)
      = 16/24 = 0.667 días

    En la práctica esta función opera en horas calendario ignorando fines
    de semana y las divide entre 24. Es una aproximación razonable dado que
    los estados en Jira pueden cambiar a cualquier hora.
    """
    if end_dt is None or start_dt is None:
        return 0.0
    if end_dt <= start_dt:
        return 0.0

    total_sec = 0.0
    cur = start_dt
    # Avanzar día a día y acumular segundos no-fin-de-semana
    while cur < end_dt:
        # Fin del día actual en la misma timezone
        day_end = datetime(cur.year, cur.month, cur.day, tzinfo=cur.tzinfo) + timedelta(days=1)
        segment_end = min(day_end, end_dt)
        if cur.weekday() < 5:   # 0=Mon ... 4=Fri
            total_sec += (segment_end - cur).total_seconds()
        cur = segment_end
    return total_sec / 86400.0   # segundos → días


def parse_changelog_metrics(changelog_json: str):
    """
    Parsea el changelog de un ticket y devuelve un dict con:
      - verified_date : date de "finalización" según VERIFIED_STATUSES_PRIORITY (o None)
      - dev_days      : días laborables desde 1ª entrada a 'In Progress'
                        hasta ÚLTIMA entrada a 'Staging' (o None)
      - qa_days       : días laborables desde 1ª entrada a 'In QA' hasta 1ª entrada
                        a cualquiera de VERIFIED_STATUSES_PRIORITY (o None)

    Todas las comparaciones de status se hacen en MAYÚSCULAS.
    """
    result = {"verified_date": None, "dev_days": None, "qa_days": None}

    if not changelog_json:
        return result
    try:
        histories = json.loads(changelog_json)
    except (json.JSONDecodeError, TypeError):
        return result

    # Construir lista ordenada de (ts, to_status_upper)
    transitions = []
    for h in histories:
        ts = parse_ts(h.get("created", ""))
        if ts is None:
            continue
        for item in h.get("items", []):
            if item.get("field") == "status":
                to_up = str(item.get("toString", "") or "").upper()
                if to_up:
                    transitions.append((ts, to_up))
    transitions.sort(key=lambda x: x[0])

    # 1ª entrada a In Progress
    first_in_progress = next(
        (ts for ts, s in transitions if s == STATUS_IN_PROGRESS), None)
    # Última entrada a Staging
    last_staging = None
    for ts, s in transitions:
        if s == STATUS_STAGING:
            last_staging = ts
    # 1ª entrada a In QA
    first_in_qa = next(
        (ts for ts, s in transitions if s == STATUS_IN_QA), None)
    # 1ª entrada a cualquiera de los estados "verified" (siguiendo prioridad)
    priority_upper = [s.upper() for s in VERIFIED_STATUSES_PRIORITY]
    first_by_status = {}
    for ts, s in transitions:
        if s in priority_upper and s not in first_by_status:
            first_by_status[s] = ts
    verified_ts = None
    for s in priority_upper:
        if s in first_by_status:
            verified_ts = first_by_status[s]
            break

    # Time in Dev = from first In Progress to last Staging
    if first_in_progress and last_staging and last_staging > first_in_progress:
        result["dev_days"] = working_days_between(first_in_progress, last_staging)

    # Time in QA = from first In QA to first Verified (priorizado)
    if first_in_qa and verified_ts and verified_ts > first_in_qa:
        result["qa_days"] = working_days_between(first_in_qa, verified_ts)

    # verified_date (para asignar período al ticket)
    if verified_ts is not None:
        result["verified_date"] = verified_ts.date()

    return result


def verified_date_from_changelog(changelog_json: str):
    """Compat: devuelve solo la fecha de verificación."""
    return parse_changelog_metrics(changelog_json)["verified_date"]


def iso_week_biweekly_label(d: date) -> str:
    """
    Etiqueta bisemanal basada en ISO week. Par (Wn, Wn+1) → 'YYW(n+1)'.
    W01 va con W02 → '26W02'. Semana impar se empareja con la siguiente.
    """
    iso_year, iso_week, _ = d.isocalendar()
    paired = iso_week if iso_week % 2 == 0 else iso_week + 1
    yy = str(iso_year)[-2:]
    return f"{yy}W{paired:02d}"


def monthly_label(d: date) -> str:
    return f"{str(d.year)[-2:]}M{d.month:02d}"


def quarterly_label(d: date) -> str:
    q = (d.month - 1) // 3 + 1
    return f"{str(d.year)[-2:]}Q{q}"


def last_n_biweekly_labels(n: int, anchor: date) -> list:
    """Devuelve las últimas n etiquetas bisemanales acabando en el ancla."""
    labels = []
    seen = set()
    d = anchor
    while len(labels) < n:
        lbl = iso_week_biweekly_label(d)
        if lbl not in seen:
            seen.add(lbl)
            labels.insert(0, lbl)
        d -= timedelta(weeks=1)
    return labels


def last_n_monthly_labels(n: int, anchor: date) -> list:
    """Devuelve las últimas n etiquetas mensuales acabando en el mes del ancla."""
    labels = []
    d = date(anchor.year, anchor.month, 1)
    for _ in range(n):
        labels.insert(0, monthly_label(d))
        d = date(d.year, d.month, 1) - timedelta(days=1)
        d = date(d.year, d.month, 1)
    return labels


def last_n_quarterly_labels(n: int, anchor: date) -> list:
    labels = []
    seen = set()
    d = anchor
    while len(labels) < n:
        lbl = quarterly_label(d)
        if lbl not in seen:
            seen.add(lbl)
            labels.insert(0, lbl)
        d -= timedelta(weeks=13)
    return labels


# ══════════════════════════════════════════════════════════════════════════════
# Parse de changelog (sólo para extraer fecha Verified for Production)
# ══════════════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════════════
# Mapping componente → (studio, game)
# ══════════════════════════════════════════════════════════════════════════════

def load_mapping(mapping_path: str) -> dict:
    """
    Lee GameStudios.csv con columnas GAME, TEAM, STUDIO.
    Devuelve {game_name_lower: {"game": ..., "studio": ...}}.
    """
    if not mapping_path:
        return {}
    comp_map = {}
    try:
        with open(mapping_path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            # Normalizar nombres de cabecera a MAYÚSCULAS
            fieldnames_norm = {fn: fn.strip().upper() for fn in reader.fieldnames or []}
            for raw in reader:
                # Re-key a MAYÚSCULAS
                row = {fieldnames_norm[k]: (v or "").strip() for k, v in raw.items() if k}
                game = row.get("GAME", "")
                studio = row.get("STUDIO", "")
                if game:
                    comp_map[game.lower()] = {"game": game, "studio": studio or "Other"}
    except FileNotFoundError:
        print(f"  Aviso: no se encontró {mapping_path}, todos los tickets irán a 'Other'")
        return {}
    return comp_map


def resolve_studio_game(components_str: str, comp_map: dict) -> tuple:
    """
    Busca en la lista de componentes del ticket cuál coincide con un juego.
    Quita el prefijo '[Game]' si está, ignora mayúsculas.
    Devuelve (studio, game).
    """
    if not components_str:
        return "Other", "Other"

    parts = [p.strip() for p in str(components_str).split(",") if p.strip()]
    for part in parts:
        # Quitar prefijo "[Game] " (con o sin espacio, case-insensitive)
        cleaned = part
        if cleaned.lower().startswith("[game]"):
            cleaned = cleaned[len("[game]"):].strip()
        key = cleaned.lower()

        if key in comp_map:
            entry = comp_map[key]
            return entry["studio"], entry["game"]

    return "Other", "Other"


# ══════════════════════════════════════════════════════════════════════════════
# Lectura de tickets del xlsx
# ══════════════════════════════════════════════════════════════════════════════

def normalize_header(h):
    return "" if h is None else str(h).strip().upper()


def load_tickets(xlsx_path: str, comp_map: dict) -> list:
    """
    Devuelve lista de dicts con:
      sp, dev_days, qa_days, verified_date, studio, game

    Time in Dev y Time in QA se calculan desde el CHANGELOG siguiendo las
    definiciones del sample oficial:
      - Time in Dev = 1ª entrada a 'In Progress' → última entrada a 'Staging'
      - Time in QA  = 1ª entrada a 'In QA' → 1ª entrada a verified/production/done
    Sólo días laborables (L–V) se cuentan.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(values_only=True))
    header_idx = {normalize_header(h): i for i, h in enumerate(header_row)}

    required = {COL_KEY, COL_STORY_POINTS, COL_COMPONENTS, COL_CHANGELOG_JSON}
    missing = [c for c in required if c not in header_idx]
    if missing:
        print(f"ERROR: faltan columnas en el xlsx: {missing}")
        print(f"Columnas encontradas: {list(header_idx.keys())}")
        sys.exit(1)

    tickets = []
    total = 0
    valid_sp = 0
    with_verified = 0
    with_dev = 0
    with_qa = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        total += 1

        sp_raw = row[header_idx[COL_STORY_POINTS]]
        try:
            sp = int(float(sp_raw))
        except (TypeError, ValueError):
            continue
        if sp not in SP_VALID:
            continue
        valid_sp += 1

        cl = row[header_idx[COL_CHANGELOG_JSON]]
        metrics = parse_changelog_metrics(str(cl) if cl else "")
        vd = metrics["verified_date"]
        if vd is None:
            continue
        with_verified += 1

        dev_days = metrics["dev_days"]  # None o float
        qa_days  = metrics["qa_days"]   # None o float
        if dev_days is not None and dev_days > 0:
            with_dev += 1
        if qa_days is not None and qa_days > 0:
            with_qa += 1

        components = row[header_idx[COL_COMPONENTS]] or ""
        studio, game = resolve_studio_game(str(components), comp_map)

        tickets.append({
            "sp":            sp,
            "dev_days":      dev_days if dev_days is not None else 0.0,
            "qa_days":       qa_days  if qa_days  is not None else 0.0,
            "verified_date": vd,
            "studio":        studio,
            "game":          game,
        })

    wb.close()
    print(f"  Filas leídas                   : {total}")
    print(f"  Con SP válido (3/5/8/13)       : {valid_sp}")
    print(f"  Con fecha de finalización      : {with_verified}")
    print(f"  Con Time in Dev calculable     : {with_dev}")
    print(f"  Con Time in QA calculable      : {with_qa}")
    return tickets


# ══════════════════════════════════════════════════════════════════════════════
# Agregación para las páginas 4 y 5
# ══════════════════════════════════════════════════════════════════════════════

def aggregate_periods(tickets: list, anchor: date) -> dict:
    bw_labels = last_n_biweekly_labels(PERIODS_BACK_BIWEEKLY, anchor)
    m_labels  = last_n_monthly_labels(PERIODS_BACK_MONTHLY, anchor)
    q_labels  = last_n_quarterly_labels(PERIODS_BACK_QUARTERLY, anchor)

    bw_vel = defaultdict(float)
    m_vel  = defaultdict(float)
    st_mon = defaultdict(lambda: defaultdict(float))
    st_qtr = defaultdict(lambda: defaultdict(float))
    gm_mon = defaultdict(lambda: defaultdict(float))
    gm_qtr = defaultdict(lambda: defaultdict(float))
    st_tot = defaultdict(float)
    gm_tot = defaultdict(float)

    for t in tickets:
        sp = t["sp"]
        vd = t["verified_date"]
        bw = iso_week_biweekly_label(vd)
        mo = monthly_label(vd)
        qt = quarterly_label(vd)

        bw_vel[bw] += sp
        m_vel[mo]  += sp

        st_mon[mo][t["studio"]] += sp
        st_qtr[qt][t["studio"]] += sp
        gm_mon[mo][t["game"]]   += sp
        gm_qtr[qt][t["game"]]   += sp
        st_tot[t["studio"]]     += sp
        gm_tot[t["game"]]       += sp

    return {
        "bw_labels": bw_labels, "m_labels": m_labels, "q_labels": q_labels,
        "bw_vel": bw_vel, "m_vel": m_vel,
        "st_mon": st_mon, "st_qtr": st_qtr,
        "gm_mon": gm_mon, "gm_qtr": gm_qtr,
        "st_tot": st_tot, "gm_tot": gm_tot,
    }


# ══════════════════════════════════════════════════════════════════════════════
# Outliers (usado en histogramas)
# ══════════════════════════════════════════════════════════════════════════════

def outlier_day_values(day_values):
    if len(day_values) < 4:
        return set()
    rounded = [int(round(v)) for v in day_values]
    s = sorted(rounded)
    q1 = s[len(s) // 4]
    q3 = s[(len(s) * 3) // 4]
    iqr = q3 - q1
    if iqr == 0:
        return set()
    return {x for x in set(rounded)
            if x > q3 + OUTLIER_IQR_MULT * iqr or x < q1 - OUTLIER_IQR_MULT * iqr}


# ══════════════════════════════════════════════════════════════════════════════
# Gráficos
# ══════════════════════════════════════════════════════════════════════════════

BG = "#EBF3FD"
GRID = "#c8d8ea"
SPINE = "#b0c4d8"


def _init_fig(figsize):
    fig, ax = plt.subplots(figsize=figsize)
    fig.patch.set_facecolor(BG)
    ax.set_facecolor(BG)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    for s_ in ["left", "bottom"]:
        ax.spines[s_].set_color(SPINE)
    ax.grid(axis="y", color=GRID, linewidth=0.5, linestyle="--", zorder=0)
    return fig, ax


def _save_fig(fig, title):
    safe = title.replace(" ", "_").replace("/", "_").replace("(", "").replace(")", "")[:50]
    path = f"/tmp/chart_{safe}.png"
    plt.tight_layout(pad=0.4)
    fig.savefig(path, dpi=130, bbox_inches="tight")
    plt.close(fig)
    return path


def make_histogram(day_values, title, figsize=(5, 3)):
    fig, ax = _init_fig(figsize)

    if not day_values:
        ax.text(0.5, 0.5, "No data", transform=ax.transAxes,
                ha="center", va="center", fontsize=10, color="#a09e98")
        ax.set_title(title, fontsize=9, pad=4, color="#1a1917")
        ax.set_xticks([])
        ax.set_yticks([])
        return _save_fig(fig, title)

    rounded = [int(round(v)) for v in day_values]
    counts = defaultdict(int)
    for d in rounded:
        counts[d] += 1
    x_vals = list(range(min(rounded), max(rounded) + 1))
    y_vals = [counts.get(x, 0) for x in x_vals]

    outliers = outlier_day_values(day_values)
    bar_colors = [COLOR_OUTLIER if x in outliers else COLOR_BAR for x in x_vals]

    ax.bar(x_vals, y_vals, color=bar_colors, width=0.7, zorder=3,
           edgecolor="white", linewidth=0.5)
    ax.set_xticks(x_vals)
    ax.set_xticklabels([str(x) for x in x_vals], fontsize=8)
    ax.yaxis.set_major_locator(mticker.MaxNLocator(integer=True))
    ax.tick_params(axis="y", labelsize=8)
    ax.set_title(title, fontsize=9, pad=4, color="#1a1917")
    ax.set_xlabel("Working days", fontsize=8, color="#6b6860")
    ax.set_ylabel("Number of tickets", fontsize=8, color="#6b6860")
    return _save_fig(fig, title)


def make_line_chart(labels, values, title, ylabel, figsize=(5.5, 3), integer_y=False):
    fig, ax = _init_fig(figsize)
    xs = range(len(labels))
    ys = [v if v is not None else float("nan") for v in values]

    ax.plot(xs, ys, color=COLOR_LINE, linewidth=2, marker="o",
            markersize=5, zorder=3)
    ax.set_xticks(list(xs))
    ax.set_xticklabels(labels, rotation=45, ha="right", fontsize=8)
    if integer_y:
        ax.yaxis.set_major_formatter(mticker.FormatStrFormatter("%.0f"))
    else:
        ax.yaxis.set_major_formatter(mticker.FormatStrFormatter("%.1f"))
    ax.tick_params(axis="y", labelsize=8)
    ax.set_title(title, fontsize=9, pad=4, color="#1a1917")
    ax.set_ylabel(ylabel, fontsize=8, color="#6b6860")
    return _save_fig(fig, title)


def make_bar_chart(labels, values, title, ylabel, figsize=(5.5, 3), integer_y=False):
    """Barras verticales — mismas dimensiones que make_line_chart."""
    fig, ax = _init_fig(figsize)
    xs = list(range(len(labels)))
    ys = [v if v is not None else 0 for v in values]

    ax.bar(xs, ys, color=COLOR_BAR, width=0.65, zorder=3,
           edgecolor="white", linewidth=0.5)
    ax.set_xticks(xs)
    ax.set_xticklabels(labels, rotation=45, ha="right", fontsize=8)
    if integer_y:
        ax.yaxis.set_major_formatter(mticker.FormatStrFormatter("%.0f"))
    else:
        ax.yaxis.set_major_formatter(mticker.FormatStrFormatter("%.1f"))
    ax.tick_params(axis="y", labelsize=8)
    ax.set_title(title, fontsize=9, pad=4, color="#1a1917")
    ax.set_ylabel(ylabel, fontsize=8, color="#6b6860")
    return _save_fig(fig, title)


def make_stacked_bar(labels, series: dict, colors_list, title, ylabel, figsize=(6, 3)):
    fig, ax = _init_fig(figsize)
    xs = list(range(len(labels)))
    bottoms = [0.0] * len(labels)
    for i, (name, vals) in enumerate(series.items()):
        col = colors_list[i % len(colors_list)]
        ax.bar(xs, vals, bottom=bottoms, color=col, label=name, width=0.6)
        bottoms = [b + v for b, v in zip(bottoms, vals)]
    ax.set_xticks(xs)
    ax.set_xticklabels(labels, rotation=45, ha="right", fontsize=8)
    ax.tick_params(axis="y", labelsize=8)
    ax.set_title(title, fontsize=9, pad=4, color="#1a1917")
    ax.set_ylabel(ylabel, fontsize=8, color="#6b6860")
    return _save_fig(fig, title)


def make_donut(labels, values, colors_list, title, figsize=(4, 3.2)):
    fig, ax = plt.subplots(figsize=figsize)
    fig.patch.set_facecolor(BG)
    total = sum(values)
    if total == 0:
        ax.text(0.5, 0.5, "No data", transform=ax.transAxes,
                ha="center", va="center", fontsize=10, color="#a09e98")
        ax.set_title(title, fontsize=9, pad=4, color="#1a1917")
        ax.axis("off")
        return _save_fig(fig, title)

    pct_labels = [f"{l}  {v/total*100:.1f}%" for l, v in zip(labels, values)]
    wedges, _ = ax.pie(
        values, labels=None, colors=colors_list[:len(values)],
        startangle=90,
        wedgeprops={"width": 0.5, "edgecolor": "white", "linewidth": 0.8}
    )
    ax.legend(wedges, pct_labels, loc="center left",
              bbox_to_anchor=(1, 0.5), fontsize=7, frameon=False)
    ax.set_title(title, fontsize=9, pad=6, color="#1a1917")
    return _save_fig(fig, title)


# ══════════════════════════════════════════════════════════════════════════════
# Construcción del PDF
# ══════════════════════════════════════════════════════════════════════════════

DARK_BLUE = colors.HexColor("#1F4E79")
TEXT_MID  = colors.HexColor("#6b6860")
TEXT_DARK = colors.HexColor("#1a1917")


def build_pdf(xlsx_tickets: list, agg: dict, target_start: date, target_end: date,
              report_month: str, output_path: str):
    W, H = A4
    doc = BaseDocTemplate(
        output_path, pagesize=A4,
        leftMargin=1.8 * cm, rightMargin=1.8 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm
    )
    frame = Frame(
        doc.leftMargin, doc.bottomMargin,
        W - doc.leftMargin - doc.rightMargin,
        H - doc.topMargin - doc.bottomMargin,
        id="main"
    )

    def header_footer(canvas, doc):
        canvas.saveState()
        canvas.setFillColor(DARK_BLUE)
        canvas.rect(0, H - 1.2 * cm, W, 1.2 * cm, fill=1, stroke=0)
        canvas.setFont("Helvetica-Bold", 9)
        canvas.setFillColor(colors.white)
        canvas.drawString(1.8 * cm, H - 0.8 * cm, "Game Server Report")
        canvas.drawRightString(W - 1.8 * cm, H - 0.8 * cm, report_month)
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(TEXT_MID)
        canvas.drawCentredString(W / 2, 0.8 * cm, f"Page {doc.page}")
        canvas.restoreState()

    doc.addPageTemplates([PageTemplate(id="main", frames=[frame], onPage=header_footer)])

    title_style = ParagraphStyle("title", fontName="Helvetica-Bold", fontSize=16,
                                  textColor=DARK_BLUE, spaceAfter=6)
    cover_title_style = ParagraphStyle("cover_title", fontName="Helvetica-Bold", fontSize=24,
                                        textColor=DARK_BLUE, alignment=TA_CENTER, spaceAfter=8)
    cover_sub_style = ParagraphStyle("cover_sub", fontName="Helvetica", fontSize=12,
                                      textColor=TEXT_MID, alignment=TA_CENTER, spaceAfter=6)
    note_style  = ParagraphStyle("note", fontName="Helvetica-Oblique", fontSize=8,
                                  textColor=TEXT_MID, spaceAfter=10, leading=11)
    intro_style = ParagraphStyle("intro", fontName="Helvetica", fontSize=10,
                                  textColor=TEXT_DARK, leading=14, alignment=TA_LEFT)

    story = []
    IMG_W = (W - doc.leftMargin - doc.rightMargin - 0.4 * cm) / 2.0
    IMG_W_FULL = W - doc.leftMargin - doc.rightMargin

    def img(path, width=None):
        from PIL import Image as PILImage
        w = width or IMG_W
        with PILImage.open(path) as im:
            iw, ih = im.size
        return Image(path, width=w, height=w * ih / iw)

    def grid_4(paths):
        t = Table(
            [[img(paths[0]), img(paths[1])],
             [img(paths[2]), img(paths[3])]],
            colWidths=[IMG_W + 0.2 * cm, IMG_W + 0.2 * cm]
        )
        t.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING",  (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING",   (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 2),
        ]))
        story.append(t)

    def two_charts(p1, p2):
        t = Table(
            [[img(p1), img(p2)]],
            colWidths=[IMG_W + 0.2 * cm, IMG_W + 0.2 * cm]
        )
        t.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING",  (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ]))
        story.append(t)

    # ── Page 1 — Cover ────────────────────────────────────────────────────────
    story.append(Spacer(1, 3 * cm))
    story.append(Paragraph("Game Server Report", cover_title_style))
    story.append(Paragraph(f"{report_month} · Metrics", cover_sub_style))
    story.append(Spacer(1, 1 * cm))
    story.append(Paragraph(
        "This report provides a comprehensive overview of the GameServer team's "
        "key performance metrics, enabling an in-depth understanding of how we are "
        "progressing against our objectives. It is organised into sections covering "
        "time spent in development, time spent in QA, team velocity, and effort "
        "distribution across studios and games.",
        intro_style
    ))
    story.append(Spacer(1, 0.4 * cm))
    story.append(Paragraph(
        f"<b>Period analysed:</b> {target_start.isoformat()} → {target_end.isoformat()}",
        intro_style
    ))
    story.append(PageBreak())

    # Ayuda para agregar tiempos por SP y período (para páginas 2 y 3)
    def build_avg_by_period(metric_key: str):
        """
        Devuelve:
          bw[sp][label] = promedio días (o None si sin datos)
          m[sp][label]  = promedio días (o None)
        Excluyendo tickets con 0 días en la métrica.
        """
        bw_vals = {sp: defaultdict(list) for sp in SP_VALID}
        m_vals  = {sp: defaultdict(list) for sp in SP_VALID}
        for t in xlsx_tickets:
            val = t[metric_key]
            if val <= 0:
                continue
            sp = t["sp"]
            vd = t["verified_date"]
            bw_vals[sp][iso_week_biweekly_label(vd)].append(val)
            m_vals[sp][monthly_label(vd)].append(val)

        def _avg(lst):
            return sum(lst) / len(lst) if lst else None

        bw_avg = {sp: {l: _avg(bw_vals[sp].get(l, [])) for l in agg["bw_labels"]}
                  for sp in SP_VALID}
        m_avg  = {sp: {l: _avg(m_vals[sp].get(l, []))  for l in agg["m_labels"]}
                  for sp in SP_VALID}
        return bw_avg, m_avg

    # ── Page 2 — Time in Development ──────────────────────────────────────────
    story.append(Paragraph("Time in Development", title_style))
    story.append(Paragraph(
        "Average working days tickets spend in development "
        "(Re-Opened + In Progress + Code Review), grouped by story points. "
        f"Ticket period determined by the 'Verified for Production' date. "
        f"Left column: last {PERIODS_BACK_BIWEEKLY} biweekly periods. "
        f"Right column: last {PERIODS_BACK_MONTHLY} months.",
        note_style
    ))

    dev_bw, dev_m = build_avg_by_period("dev_days")
    for sp in [3, 5, 8, 13]:
        bw_vals = [dev_bw[sp].get(l) for l in agg["bw_labels"]]
        m_vals  = [dev_m[sp].get(l)  for l in agg["m_labels"]]
        p_bw = make_bar_chart(agg["bw_labels"], bw_vals,
                              f"Time in Dev – {sp} SP / 2 weeks",
                              "Working days (avg)", figsize=(4.8, 2.3))
        p_m  = make_bar_chart(agg["m_labels"], m_vals,
                              f"Time in Dev – {sp} SP / Month",
                              "Working days (avg)", figsize=(4.8, 2.3))
        two_charts(p_bw, p_m)
    story.append(PageBreak())

    # ── Page 3 — Time in QA ───────────────────────────────────────────────────
    story.append(Paragraph("Time in QA", title_style))
    story.append(Paragraph(
        "Average working days tickets spend in QA "
        "(Staging + In QA), grouped by story points. "
        f"Ticket period determined by the 'Verified for Production' date. "
        f"Left column: last {PERIODS_BACK_BIWEEKLY} biweekly periods. "
        f"Right column: last {PERIODS_BACK_MONTHLY} months.",
        note_style
    ))

    qa_bw, qa_m = build_avg_by_period("qa_days")
    for sp in [3, 5, 8, 13]:
        bw_vals = [qa_bw[sp].get(l) for l in agg["bw_labels"]]
        m_vals  = [qa_m[sp].get(l)  for l in agg["m_labels"]]
        p_bw = make_bar_chart(agg["bw_labels"], bw_vals,
                              f"Time in QA – {sp} SP / 2 weeks",
                              "Working days (avg)", figsize=(4.8, 2.3))
        p_m  = make_bar_chart(agg["m_labels"], m_vals,
                              f"Time in QA – {sp} SP / Month",
                              "Working days (avg)", figsize=(4.8, 2.3))
        two_charts(p_bw, p_m)
    story.append(PageBreak())

    # ── Página 4 — Team Velocity ──────────────────────────────────────────────
    story.append(Paragraph("Team Velocity", title_style))
    story.append(Paragraph(
        "Sum of story points of tickets that reached 'Verified for Production'. "
        f"Last {PERIODS_BACK_BIWEEKLY} biweekly periods (left) and "
        f"{PERIODS_BACK_MONTHLY} months (right). Anchored to the last completed month.",
        note_style
    ))
    bw_vel_vals = [agg["bw_vel"].get(l, 0) for l in agg["bw_labels"]]
    mo_vel_vals = [agg["m_vel"].get(l, 0)  for l in agg["m_labels"]]
    p_bw = make_line_chart(agg["bw_labels"], bw_vel_vals,
                           "Velocity per 2 Weeks (To Verified for Production)",
                           "Num Story Points", integer_y=True)
    p_mo = make_line_chart(agg["m_labels"], mo_vel_vals,
                           "Velocity per Month (To Verified for Production)",
                           "Num Story Points", integer_y=True)
    two_charts(p_bw, p_mo)
    story.append(PageBreak())

    # ── Page 5 — Effort Distribution ──────────────────────────────────────────
    story.append(Paragraph("Effort Distribution", title_style))
    story.append(Paragraph(
        "Story points released per studio and per game. "
        f"Last {PERIODS_BACK_MONTHLY} months and {PERIODS_BACK_QUARTERLY} quarters.",
        note_style
    ))

    studios = sorted(agg["st_tot"], key=lambda s: -agg["st_tot"][s])
    games   = sorted(agg["gm_tot"], key=lambda g: -agg["gm_tot"][g])[:10]

    # Donut global por estudio
    if studios:
        vals = [agg["st_tot"][s] for s in studios]
        donut_path = make_donut(studios, vals, STUDIO_COLORS,
                                 "Effort per Studio (Global)", figsize=(5, 3))
        story.append(img(donut_path, IMG_W_FULL * 0.65))
        story.append(Spacer(1, 0.25 * cm))

    # Estudios: mensual + trimestral
    if studios:
        st_mon_series = {s: [agg["st_mon"].get(l, {}).get(s, 0) for l in agg["m_labels"]]
                         for s in studios}
        st_qtr_series = {s: [agg["st_qtr"].get(l, {}).get(s, 0) for l in agg["q_labels"]]
                         for s in studios}
        p1 = make_stacked_bar(agg["m_labels"], st_mon_series, STUDIO_COLORS,
                              "Effort per Studio – Monthly", "Story Points")
        p2 = make_stacked_bar(agg["q_labels"], st_qtr_series, STUDIO_COLORS,
                              "Effort per Studio – Quarterly", "Story Points")
        two_charts(p1, p2)
        story.append(Spacer(1, 0.15 * cm))

    # Juegos: mensual + trimestral
    if games:
        gm_mon_series = {g: [agg["gm_mon"].get(l, {}).get(g, 0) for l in agg["m_labels"]]
                         for g in games}
        gm_qtr_series = {g: [agg["gm_qtr"].get(l, {}).get(g, 0) for l in agg["q_labels"]]
                         for g in games}
        p1 = make_stacked_bar(agg["m_labels"], gm_mon_series, GAME_COLORS,
                              "Effort per Game – Monthly", "Story Points")
        p2 = make_stacked_bar(agg["q_labels"], gm_qtr_series, GAME_COLORS,
                              "Effort per Game – Quarterly", "Story Points")
        two_charts(p1, p2)

    doc.build(story)
    print(f"  PDF generado → {output_path}")


# ══════════════════════════════════════════════════════════════════════════════
# CLI
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Genera el PDF mensual de performance del CTO.")
    parser.add_argument("--xlsx", required=True, help="Ruta al GS_jira_year_report.xlsx")
    parser.add_argument("--mapping", default="GameStudios.csv",
                        help="Ruta al CSV de mapeo (default: GameStudios.csv)")
    parser.add_argument("--output", default="GS_CTO_Report.pdf", help="Nombre del PDF de salida")
    parser.add_argument("--month", default=None,
                        help="Mes del informe, YYYY-MM (default: mes anterior al actual)")
    args = parser.parse_args()

    # Resolver mes objetivo
    if args.month:
        try:
            dt = datetime.strptime(args.month, "%Y-%m")
            target_start = date(dt.year, dt.month, 1)
            next_month = dt.month % 12 + 1
            next_year  = dt.year + (1 if dt.month == 12 else 0)
            target_end = date(next_year, next_month, 1) - timedelta(days=1)
            report_month = dt.strftime("%B %Y")
        except ValueError:
            print(f"ERROR: formato de --month inválido: {args.month} (usa YYYY-MM)")
            sys.exit(1)
    else:
        target_start, target_end, report_month = previous_month_bounds()

    # El ancla para páginas 4/5 es el último día del mes objetivo
    anchor = target_end

    print(f"\nGS CTO Report Generator")
    print(f"  Mes del informe : {report_month}")
    print(f"  Rango de fechas : {target_start} → {target_end}")
    print(f"  XLSX            : {args.xlsx}")
    print(f"  Mapping         : {args.mapping}")
    print(f"  PDF de salida   : {args.output}\n")

    print("Cargando mapping de estudios/juegos...")
    comp_map = load_mapping(args.mapping)
    print(f"  {len(comp_map)} mapeos cargados")

    print("Cargando tickets...")
    tickets = load_tickets(args.xlsx, comp_map)

    print("Agregando datos por período...")
    agg = aggregate_periods(tickets, anchor)
    print(f"  Labels bisemanales : {agg['bw_labels']}")
    print(f"  Labels mensuales   : {agg['m_labels']}")
    print(f"  Labels trimestrales: {agg['q_labels']}")

    print("\nGenerando PDF...")
    build_pdf(tickets, agg, target_start, target_end, report_month, args.output)

    print("\nListo.")


if __name__ == "__main__":
    main()
